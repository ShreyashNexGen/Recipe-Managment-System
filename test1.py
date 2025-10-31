from flask import (
    Flask,
    render_template,
    jsonify,
    request,
    redirect,
    url_for,
    session,
    flash,
)
import pandas as pd
from io import BytesIO
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
import json
import logging
import datetime
import webbrowser
from datetime import datetime, timedelta
from flask_socketio import SocketIO, emit
from opcua import Client, ua
from flask import send_from_directory
import threading
from threading import Thread
import os
import socket
from flask import send_file
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import time
import requests
import json
import sqlite3
import platform
import datetime
import subprocess
import wmi  # For Windows systems (install with `pip install WMI`)
from bson import json_util
from pymongo import MongoClient
from flask_cors import CORS
import pyodbc
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import pandas as pd
import json
from openpyxl import Workbook
from flask import send_file
import io
import openpyxl
from openpyxl.styles import Alignment
from io import BytesIO
from flask import send_file, flash, redirect, url_for

app = Flask(__name__)
CORS(app)  # Enable CORS for frontend communication
# socketio = SocketIO(app, cors_allowed_origins="*", async_mode='gevent')
# OPC UA connection options
# Connection details
# server = 'DESKTOP-9G39B01\WINCC'
# database = 'A2Z_DB'3
server = "SHREYASHNEXGEN\WINCCFLEX2014"
database = "Shreyash"
conn = pyodbc.connect(
    f"DRIVER={{ODBC Driver 17 for SQL Server}};"
    f"SERVER={server};"
    f"DATABASE={database};"
    f"Trusted_Connection=yes;"
)
import os


def fetch_endpoint_url():
    """Fetch plcIp and plcPort from MSSQL database and construct ENDPOINT_URL."""
    try:
        cursor = conn.cursor()

        # Fetch plcIp and plcPort from the database
        cursor.execute("SELECT plcIp, plcPort FROM Plc_Table")
        result = cursor.fetchone()

        if result:
            plc_ip, plc_port = result
            endpoint_url = f"opc.tcp://{plc_ip}:{plc_port}"
            return endpoint_url
        else:
            return "No PLC data available in the database."
    except pyodbc.Error as e:
        return f"Database error: {e}"
    finally:
        cursor.close()


ENDPOINT_URL = fetch_endpoint_url()
app.secret_key = "your_secret_key"
app.config["SESSION_TYPE"] = "filesystem"
app.permanent_session_lifetime = timedelta(minutes=600)  # Session timeout


def init_db():
    """Initialize the MSSQL database."""
    try:
        cursor = conn.cursor()

        # Create tables
        tables = {
            "Tag_Table": """
                CREATE TABLE Tag_Table (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    tagId NVARCHAR(255) UNIQUE NOT NULL,
                    tagName NVARCHAR(255) UNIQUE NOT NULL,
                    tagAddress NVARCHAR(255) NOT NULL,
                    plcId NVARCHAR(255) NOT NULL
                )
            """,
            "Raw_Materials": """
                    CREATE TABLE Raw_Materials (
                        material_Id INT IDENTITY(1,1) PRIMARY KEY,
                        type NVARCHAR(255) NOT NULL,
                        width NVARCHAR(255) NOT NULL,
                        part_no NVARCHAR(255) NOT NULL,
                        materialType NVARCHAR(255) NOT NULL UNIQUE,
                        make NVARCHAR(255) NOT NULL,
                        [user] NVARCHAR(255) NOT NULL,
                        barcode NVARCHAR(255) NOT NULL
                    )
            """,
            "Alu_Raw_Materials": """
                    CREATE TABLE Alu_Raw_Materials (
                        material_Id INT IDENTITY(1,1) PRIMARY KEY,
                        thickness NVARCHAR(255) NOT NULL,
                        width NVARCHAR(255) NOT NULL,
                        part_no NVARCHAR(255) NOT NULL,
                        materialType NVARCHAR(255) NOT NULL UNIQUE,
                        make NVARCHAR(255) NOT NULL,
                        [user] NVARCHAR(255) NOT NULL,
                        barcode NVARCHAR(255) NOT NULL
                    )
            """,
            "h_Raw_Materials": """
                    CREATE TABLE h_Raw_Materials (
                        material_Id INT IDENTITY(1,1) PRIMARY KEY,
                        length NVARCHAR(255) NOT NULL,
                        width NVARCHAR(255) NOT NULL,
                        height NVARCHAR(255) NOT NULL,
                        part_no NVARCHAR(255) NOT NULL,
                        materialType NVARCHAR(255) NOT NULL UNIQUE,
                        make NVARCHAR(255) NOT NULL,
                        [user] NVARCHAR(255) NOT NULL,
                        barcode NVARCHAR(255) NOT NULL
                    )
            """,
            "Recipe": """ 
                     CREATE TABLE Recipe (
                       Recipe_ID INT PRIMARY KEY,
                        Filter_Size NVARCHAR(255),
                        Filter_Code NVARCHAR(255),
                        Art_No INT,
                        Recipe_Name NVARCHAR(255)
                    )
            """,
            "Recipe_Details1": """
                CREATE TABLE Recipe_Details1 (
                        Id INT IDENTITY(1,1) PRIMARY KEY,
                        Pos1 NVARCHAR(255),
                        Pos2 NVARCHAR(255),
                        Pos3 NVARCHAR(255),
                        Pos4 NVARCHAR(255),
                        Pos5 NVARCHAR(255),
                        Pos6 NVARCHAR(255),
                        Pos7 NVARCHAR(255),
                        Pos8 NVARCHAR(255),
                        Pos9 NVARCHAR(255),
                        Alu_coil_width NVARCHAR(255),
                        Alu_roller_type INT,
                        Spacer FLOAT,
                        Alu_Material NVARCHAR(255),
                        House_Material NVARCHAR(255),
                        Recipe_ID INT FOREIGN KEY REFERENCES Recipe(Recipe_ID)
                    )
            """,
            "Recipe_Log": """
    CREATE TABLE Recipe_Log (
        Batch_No INT IDENTITY(1,1) NOT NULL PRIMARY KEY,
        Batch_Code NVARCHAR(255) NOT NULL,
        Timestamp DATETIME2(7) NOT NULL DEFAULT (GETDATE()),
        SerialNo NVARCHAR(255),
        NgStatus NVARCHAR(255),
        Avg_Air_Flow NVARCHAR(255),
        Avg_Result NVARCHAR(255),
        Batch_Completion_Status NVARCHAR(255) NOT NULL,
        Recipe_Id INT,
        Recipe_Name NVARCHAR(255) NOT NULL,
        Filter_Size NVARCHAR(255),
        Filter_Code NVARCHAR(255),
        Art_No NVARCHAR(255),
        Pos1 NVARCHAR(255),
        Pos2 NVARCHAR(255),
        Pos3 NVARCHAR(255),
        Pos4 NVARCHAR(255),
        Pos5 NVARCHAR(255),
        Pos6 NVARCHAR(255),
        Pos7 NVARCHAR(255),
        Pos8 NVARCHAR(255),
        Pos9 NVARCHAR(255),
        Alu_Material NVARCHAR(255),
        House_Material NVARCHAR(255),
        Pleat_Height NVARCHAR(255),
        Soft_Touch NVARCHAR(255),
        Feeder1_Media_Thickness NVARCHAR(255),
        Feeder1_Park_Position NVARCHAR(255),
        Foil1_Length NVARCHAR(255),
        Foil1_Length_Offset NVARCHAR(255),
        Foil1_Width NVARCHAR(255),
        Puller_Start_Position NVARCHAR(255),
        Autofeeding_Offset NVARCHAR(255),
        Filter_Box_Height NVARCHAR(255),
        Filter_Box_Length NVARCHAR(255),
        Pleat_Pitch NVARCHAR(255),
        Pleat_Counts NVARCHAR(255),
        Foil_Low_Diameter NVARCHAR(255),
        Feeding_Conveyor_Speed NVARCHAR(255),
        Pack_Transfer_Rev_Position NVARCHAR(255),
        Foil1_Tension_Set_Point NVARCHAR(255),
        Foil2_Tension_Set_Point NVARCHAR(255),
        Blade_Opening NVARCHAR(255),
        Press_Touch NVARCHAR(255),
        Feeder2_Media_Thickness NVARCHAR(255),
        Feeder2_Park_Position NVARCHAR(255),
        Foil2_Length NVARCHAR(255),
        Foil2_Length_Offset NVARCHAR(255),
        Foil2_Width NVARCHAR(255),
        Puller_End_Position NVARCHAR(255),
        Puller2_Feed_Correction NVARCHAR(255),
        Puller_Extra_Stroke_Enable NVARCHAR(255),
        Filter_Box_Width NVARCHAR(255),
        Lid_Placement_Enable NVARCHAR(255),
        Lid_Placement_Position NVARCHAR(255),
        Sync_Table_Start_Position NVARCHAR(255),
        Batch_Count NVARCHAR(255),
        Discharge_Conveyor_Speed NVARCHAR(255),
        Pack_Transfer_Park_Position NVARCHAR(255),
        Media_Tension_Set_Point NVARCHAR(255),
        DatabaseAvailable NVARCHAR(255),
        Width NVARCHAR(255),
        Height NVARCHAR(255),
        Depth NVARCHAR(255),
        Inspection_Art_No NVARCHAR(255),
        Air_Flow_Set NVARCHAR(255),
        Pressure_Drop_Setpoint NVARCHAR(255),
        Lower_Tolerance1 NVARCHAR(255),
        Lower_Tolerance2 NVARCHAR(255),
        Upper_Tolerance1 NVARCHAR(255),
        Upper_Tolerance2 NVARCHAR(255),
        Upper_fan_Speed NVARCHAR(255),
        Lower_fan_Speed NVARCHAR(255),
        Airflow_upper_limit NVARCHAR(255),
        Airflow_lower_limit NVARCHAR(255),
        Average_motor_RPM NVARCHAR(255),
        testing_duration NVARCHAR(255),
        Average_Pressure NVARCHAR(255),
        user_Id NVARCHAR(50)
    );


            """,
            "Sub_Menu": """
    CREATE TABLE Sub_Menu (
        Recipe_ID INT NOT NULL PRIMARY KEY,
        Pleat_Height DECIMAL(10, 2),
        Soft_Touch DECIMAL(10, 2),
        Feeder1_Media_Thickness DECIMAL(10, 2),
        Feeder1_Park_Position DECIMAL(10, 2),
        Foil1_Length DECIMAL(10, 2),
        Foil1_Length_Offset DECIMAL(10, 2),
        Foil1_Width DECIMAL(10, 2),
        Puller_Start_Position DECIMAL(10, 2),
        Autofeeding_Offset DECIMAL(10, 2),
        Filter_Box_Height DECIMAL(10, 2),
        Filter_Box_Length DECIMAL(10, 2),
        Pleat_Pitch DECIMAL(10, 2),
        Pleat_Counts INT,
        Foil_Low_Diameter DECIMAL(10, 2),
        Feeding_Conveyor_Speed DECIMAL(10, 2),
        Pack_Transfer_Rev_Position DECIMAL(10, 2),
        Foil1_Tension_Set_Point DECIMAL(10, 2),
        Foil2_Tension_Set_Point DECIMAL(10, 2),
        Blade_Opening DECIMAL(10, 2),
        Press_Touch DECIMAL(10, 2),
        Feeder2_Media_Thickness DECIMAL(10, 2),
        Feeder2_Park_Position DECIMAL(10, 2),
        Foil2_Length DECIMAL(10, 2),
        Foil2_Length_Offset DECIMAL(10, 2),
        Foil2_Width DECIMAL(10, 2),
        Puller_End_Position DECIMAL(10, 2),
        Puller2_Feed_Correction DECIMAL(10, 2),
        Puller_Extra_Stroke_Enable BIT,
        Filter_Box_Width DECIMAL(10, 2),
        Lid_Placement_Enable BIT,
        Lid_Placement_Position DECIMAL(10, 2),
        Sync_Table_Start_Position DECIMAL(10, 2),
        Batch_Count INT,
        Discharge_Conveyor_Speed DECIMAL(10, 2),
        Pack_Transfer_Park_Position DECIMAL(10, 2),
        Media_Tension_Set_Point DECIMAL(10, 2)
    );
""",
            "Inspection_Settings": """

                 CREATE TABLE Inspection_Settings (
                     Inspection_ID INT PRIMARY KEY IDENTITY(1,1),
    Recipe_ID INT FOREIGN KEY REFERENCES Recipe(Recipe_ID),
    databaseAvailable NVARCHAR(255),
    Width NVARCHAR(255),
    Height NVARCHAR(255),
    Depth NVARCHAR(255),
    Art_No NVARCHAR(255),
    Air_Flow_Set NVARCHAR(255),
    Pressure_Drop_Setpoint NVARCHAR(255),
    Lower_Tolerance1 NVARCHAR(255),
    Lower_Tolerance2 NVARCHAR(255),
          Upper_Tolerance1 NVARCHAR(255),
    Upper_Tolerance2 NVARCHAR(255)
    );

                """,
            "Live_Tags": """
                CREATE TABLE Live_Tags (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    tagId NVARCHAR(255) NOT NULL,
                    value NVARCHAR(255) NOT NULL,
                    timestamp DATETIME DEFAULT GETDATE(),
                    FOREIGN KEY (tagId) REFERENCES Tag_Table(tagId)
                )
            """,
            "Live_Log": """
                CREATE TABLE Live_Log (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    tagName NVARCHAR(255) NOT NULL,
                    tagId NVARCHAR(255) NOT NULL,
                    value NVARCHAR(255) NOT NULL,
                    timestamp DATETIME DEFAULT GETDATE()
                )
            """,
            "connection_status": """
                CREATE TABLE connection_status (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    internet_status NVARCHAR(50) NOT NULL,
                    plc_status NVARCHAR(50) NOT NULL,
                    timestamp DATETIME DEFAULT GETDATE()
                )
            """,
            "internet_status": """
                CREATE TABLE internet_status (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    startTime NVARCHAR(50) NOT NULL,
                    endTime NVARCHAR(50) NOT NULL,
                    time_Duration NVARCHAR(50) NOT NULL,
                    status NVARCHAR(50) NOT NULL
                )
            """,
            "plc_status": """
                CREATE TABLE plc_status (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    startTime NVARCHAR(50) NOT NULL,
                    endTime NVARCHAR(50) NOT NULL,
                    time_Duration NVARCHAR(50) NOT NULL,
                    status NVARCHAR(50) NOT NULL
                )
            """,
            "users": """
                CREATE TABLE users (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    username NVARCHAR(255) NOT NULL UNIQUE,
                    email NVARCHAR(255) NOT NULL UNIQUE,
                    password NVARCHAR(255) NOT NULL,
                    is_admin INT DEFAULT 0,
                    roles NVARCHAR(MAX)
                )
            """,
            "user_activity": """
                CREATE TABLE user_activity (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    username NVARCHAR(255) NOT NULL,
                    login_time DATETIME,
                    logout_time DATETIME,
                    FOREIGN KEY (username) REFERENCES users(username)
                )
            """,
            "Plc_Table": """
                CREATE TABLE Plc_Table (
                    Id INT IDENTITY(1,1) PRIMARY KEY,
                    plcId INT DEFAULT 1,
                    plcName NVARCHAR(50) DEFAULT 'Plc1',
                    plcIp NVARCHAR(50) DEFAULT '192.168.0.1',
                    plcPort INT DEFAULT 4840,
                    intervalTime INT DEFAULT 20,
                    serial_Key NVARCHAR(50) DEFAULT '12345'
                )
            """,
            "BatchTracker": """
                CREATE TABLE BatchTracker (
                    id INT IDENTITY(1,1) PRIMARY KEY,
                    number_of_batches_created INT NOT NULL
                )
            """,
        }

        for table, query in tables.items():
            cursor.execute(
                f"IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='{table}' AND xtype='U') {query}"
            )
            conn.commit()

        # Insert default Plc_Table entry if table is empty
        cursor.execute("SELECT COUNT(*) FROM Plc_Table")
        if cursor.fetchone()[0] == 0:
            cursor.execute(
                """
                INSERT INTO Plc_Table (plcId, plcName, plcIp, plcPort, intervalTime, serial_Key)
                VALUES (1, 'Plc1', '192.168.0.1', 4840, 20, '11223344')
            """
            )
            conn.commit()

        # Insert default BatchTracker entry if table is empty
        cursor.execute("SELECT COUNT(*) FROM BatchTracker")
        if cursor.fetchone()[0] == 0:
            cursor.execute(
                "INSERT INTO BatchTracker (number_of_batches_created) VALUES (0)"
            )
            conn.commit()

        print("Database initialization and setup completed successfully.")
    except pyodbc.Error as e:
        print(f"Error initializing the database: {e}")


# Initialize the database
init_db()

# Hardcoded user data for simplicity
USER_DATA = {
    "username": "admin",
    "password_hash": generate_password_hash("password123"),  # Secure hashed password
}


def get_opcua_client():
    """Create and return an OPC UA client."""
    client = Client(ENDPOINT_URL)
    try:
        client.connect()
        print("Connected to OPC UA Server")
    except Exception as e:
        print(f"Failed to connect to OPC UA Server: {e}")
        return None
    return client


@app.route("/production")
def filter_log():
    if "username" not in session:
        flash("Please log in to access the dashboard.", "warning")
        return redirect(url_for("login"))
    return render_template("production.html", username=session["username"])


@app.route("/status", methods=["GET"])
def get_status():
    try:
        cursor = conn.cursor()

        # Fetch PLC IP and Interval Time
        cursor.execute("SELECT plcIp, intervalTime FROM Plc_Table")
        plc_data = cursor.fetchone()
        plc_ip = plc_data[0] if plc_data else "Not Available"
        interval_time = plc_data[1] if plc_data else "Not Set"

        # Fetch Tag Count
        cursor.execute("SELECT COUNT(*) FROM Tag_Table")
        tags_count = cursor.fetchone()[0]

        # Fetch Log Count from BatchTracker
        cursor.execute("SELECT number_of_batches_created FROM BatchTracker")
        result = cursor.fetchone()
        logs_count = result[0] if result else 0

        # Get Serial Key
        serial_key = platform.node()  # Fetches the computer's hostname as serial key

        # Update Serial Key in Plc_Table
        cursor.execute(
            "UPDATE Plc_Table SET serial_Key = ? WHERE plcId = 1", (serial_key,)
        )
        conn.commit()
        cursor.execute(
            """
         SELECT recipe_name 
        FROM recipe 
         WHERE recipe_id = (
        SELECT TOP 1 recipe_id 
        FROM recipe 
        ORDER BY recipe_id DESC
        )
        """
        )
        result = cursor.fetchone()
        last_Recipe = result[0] if result else "N/A"

        conn.commit()
        # Internet and PLC connection statuses
        internet_connected = check_internet_connection()
        plc_connected = check_plc_connection(plc_ip)

        # Fetch Recipe Name from PLC
        client = get_opcua_client()
        if client:
            recipe_name_field_path = 'ns=3;s="dbRecipe1"."Recipe"."RecipeName"'  # Example path, adjust as per your actual path
            try:
                recipe_name = client.get_node(recipe_name_field_path).get_value()
            except Exception as e:
                recipe_name = "Error Retrieving"
            client.disconnect()
        else:
            recipe_name = "Not Connected to OPC UA Server"

        # Return the statuses
        return jsonify(
            {
                "Plc_IP": plc_ip,
                "Internet_Connected": internet_connected,
                "Plc_Connected": plc_connected,
                "No_Of_Tags_Created": tags_count,
                "No_Of_Logs_Created": logs_count,
                "Interval_Time_Of_Log_Entry": interval_time,
                "Serial_Key": serial_key,
                "Recipe_Name": recipe_name,
                "Last_Recipe": last_Recipe,
            }
        )
    except pyodbc.Error as e:
        return jsonify({"error": f"Database error: {str(e)}"})
    except Exception as e:
        return jsonify({"error": str(e)})
    finally:
        cursor.close()


def check_internet_connection():
    """Check if the server has internet connectivity."""
    try:
        import socket

        socket.create_connection(("8.8.8.8", 53), timeout=2)
        return "Connected"
    except Exception:
        return "Not Connected"


def check_plc_connection(plc_ip):
    """Check if the PLC is reachable."""
    try:
        import os

        response = os.system(
            f"ping -n 1 {plc_ip}"
        )  # Replace `-c` with `-n` for Windows
        return "Connected" if response == 0 else "Not Connected"
    except Exception:
        return "Unknown"


def get_last_status(table_name):
    """Retrieve the last status from the specified table."""
    cursor = conn.cursor()
    cursor.execute(f"SELECT TOP 1 status FROM {table_name} ORDER BY id DESC")
    result = cursor.fetchone()
    cursor.close()  # Close cursor instead of closing the connection
    return result[0] if result else None


def get_last_combined_status():
    """Retrieve the last combined status (internet and PLC) from the connection_status table."""
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT TOP 1 internet_status, plc_status 
        FROM connection_status
        ORDER BY id DESC
    """
    )
    result = cursor.fetchone()
    cursor.close()  # Close cursor instead of closing the connection
    return result if result else (None, None)


def calculate_duration(start_time, end_time):
    """Calculate the duration between two timestamps."""
    start = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
    end = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")
    duration = end - start
    return str(duration)


def log_status():
    """Log statuses for Internet and PLC into tables."""
    # Initialize previous statuses and start times
    last_internet_status = check_internet_connection()
    last_plc_status = check_plc_connection("192.168.0.1")  # Replace with actual PLC IP
    internet_start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    plc_start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    while True:
        try:
            # Get current statuses
            internet_status = check_internet_connection()
            plc_status = check_plc_connection(
                "192.168.0.1"
            )  # Replace with actual PLC IP
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Get the last logged combined status
            last_combined_status = get_last_combined_status()

            # Use single connection for all operations
            cursor = conn.cursor()

            # Log to connection_status table only if status has changed
            if (internet_status, plc_status) != last_combined_status:
                cursor.execute(
                    """
                    INSERT INTO connection_status (timestamp, internet_status, plc_status)
                    VALUES (?, ?, ?)
                """,
                    (current_time, internet_status, plc_status),
                )
                conn.commit()

            # Handle Internet status changes
            if internet_status != last_internet_status:
                last_record_status = get_last_status("internet_status")
                if last_record_status != internet_status:
                    duration = calculate_duration(internet_start_time, current_time)
                    cursor.execute(
                        """
                        INSERT INTO internet_status (startTime, endTime, time_Duration, status)
                        VALUES (?, ?, ?, ?)
                    """,
                        (internet_start_time, current_time, duration, internet_status),
                    )
                    conn.commit()

                # Update start time and last status
                internet_start_time = current_time
                last_internet_status = internet_status

            # Handle PLC status changes
            if plc_status != last_plc_status:
                last_record_status = get_last_status("plc_status")
                if last_record_status != plc_status:
                    duration = calculate_duration(plc_start_time, current_time)
                    cursor.execute(
                        """
                        INSERT INTO plc_status (startTime, endTime, time_Duration, status)
                        VALUES (?, ?, ?, ?)
                    """,
                        (plc_start_time, current_time, duration, plc_status),
                    )
                    conn.commit()

                # Update start time and last status
                plc_start_time = current_time
                last_plc_status = plc_status

            # Wait for 1 second before the next check
            time.sleep(30)

        except Exception as e:
            print(f"Error: {e}")


@app.route("/status1", methods=["GET"])
def get_status_summary():
    """Fetch summary for pie charts."""
    # conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT 
            SUM(CASE WHEN internet_status = 'Connected' THEN 1 ELSE 0 END) AS internet_uptime,
            SUM(CASE WHEN internet_status = 'Not Connected' THEN 1 ELSE 0 END) AS internet_downtime,
            SUM(CASE WHEN plc_status = 'Connected' THEN 1 ELSE 0 END) AS plc_connected,
            SUM(CASE WHEN plc_status = 'Not Connected' THEN 1 ELSE 0 END) AS plc_disconnected
        FROM connection_status
    """
    )
    result = cursor.fetchone()
    # conn.close()

    return jsonify(
        {
            "internet": {"uptime": result[0], "downtime": result[1]},
            "plc": {"connected": result[2], "disconnected": result[3]},
        }
    )


def get_db_connection():
    """Establishes a connection to MSSQL Server."""
    try:
        conn = pyodbc.connect(
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={server};"
            f"DATABASE={database};"
            f"Trusted_Connection=yes;"
        )
        return conn
    except pyodbc.Error as e:
        flash(f"Database connection error: {e}", "danger")
        return None


@app.route("/")
def index():
    """Main dashboard route with pagination."""
    if "username" not in session:  # User must be logged in
        return redirect(url_for("login"))

    # Pagination
    page = max(1, request.args.get("page", 1, type=int))
    per_page = 10
    offset = (page - 1) * per_page

    conn = get_db_connection()
    if not conn:
        return redirect(url_for("login"))

    cursor = conn.cursor()

    # Fetch paginated data
    cursor.execute(
        """
     SELECT * FROM Recipe
    ORDER BY Recipe_ID  
    OFFSET ? ROWS
    FETCH NEXT ? ROWS ONLY
    """,
        (offset, per_page),
    )

    recipes = cursor.fetchall()

    # Calculate total pages
    cursor.execute("SELECT COUNT(*) FROM Recipe")
    total_count = cursor.fetchone()[0]
    total_pages = (total_count + per_page - 1) // per_page

    conn.close()

    return render_template(
        "dashboard.html",
        # recipes=recipes,  # Pass data to template
        # page=page,
        # total_pages=total_pages
    )


@app.before_request
def require_login():
    allowed_routes = [
        "login",
        "register",
        "static",
    ]  # Allow public routes and static files
    if request.endpoint not in allowed_routes and "id" not in session:
        if not request.endpoint or request.endpoint.startswith(
            "static"
        ):  # Allow static files
            return
        flash("Please log in to access this page.", "warning")
        return redirect(url_for("login"))


@app.route("/dashboard")
def dashboard():
    if "username" not in session:
        flash("Please log in to access the dashboard.", "warning")
        return redirect(url_for("login"))
    return render_template("dashboard.html", username=session["username"])


@app.route("/recipe")
def recipe_list():
    """Fetch and display paginated recipe list from MSSQL."""
    if "username" not in session:  # Check if the user is logged in
        return redirect(url_for("login"))

    # Pagination logic
    page = request.args.get("page", 1, type=int)
    per_page = 5
    offset = (page - 1) * per_page

    conn = get_db_connection()
    if not conn:
        return redirect(url_for("login"))  # Redirect if DB connection fails

    cursor = conn.cursor()

    # Fetch paginated recipes
    cursor.execute(
        """
        SELECT * FROM Recipe 
        ORDER BY Recipe_ID
        OFFSET ? ROWS 
        FETCH NEXT ? ROWS ONLY
    """,
        (offset, per_page),
    )
    recipes = cursor.fetchall()

    # Fetch distinct material types
    cursor.execute("SELECT DISTINCT materialType FROM Raw_Materials")
    pos_values = [row[0] for row in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT materialType FROM Alu_Raw_Materials")
    alu_values = [row[0] for row in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT materialType FROM h_Raw_Materials")
    h_values = [row[0] for row in cursor.fetchall()]

    # Get total recipe count
    cursor.execute("SELECT COUNT(*) FROM Recipe")
    total_count = cursor.fetchone()[0]
    total_pages = (total_count + per_page - 1) // per_page

    # ✅ Fetch last Recipe_ID
    cursor.execute("SELECT ISNULL(MAX(Recipe_ID), 0) FROM Recipe")
    last_recipe_id = cursor.fetchone()[0]
    next_recipe_id = last_recipe_id + 1  # default next ID

    conn.close()

    return render_template(
        "recipe.html",
        recipes=recipes,
        pos_values=pos_values,
        alu_values=alu_values,
        h_values=h_values,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        next_recipe_id=next_recipe_id,  # ✅ pass to template
    )


@app.route("/api/recipe_log", methods=["GET"])
def get_recipe_log():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Date filter support
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    today = datetime.now()

    # Default end_date = today
    if end_date and end_date.strip() != "":
        end_date = datetime.strptime(end_date, "%Y-%m-%dT%H:%M")
    else:
        end_date = today

    # Default start_date = yesterday
    if start_date and start_date.strip() != "":
        start_date = datetime.strptime(start_date, "%Y-%m-%dT%H:%M")
    else:
        start_date = today - timedelta(days=1)

    # Convert to string for SQL Server
    start_date_str = start_date.strftime("%Y-%m-%d %H:%M:%S")
    end_date_str = end_date.strftime("%Y-%m-%d %H:%M:%S")

    query = """
        SELECT 
            SerialNo,
            Timestamp,
            Art_No,
            Recipe_Name,
            Filter_Code,
            NgStatus,
            Upper_fan_Speed,
            Lower_fan_Speed,
            Upper_Tolerance1,
            Lower_Tolerance1,
            Airflow_upper_limit,
            Airflow_lower_limit,
            Upper_Tolerance2,
            Lower_Tolerance2,
            Average_motor_RPM,
            Avg_Air_Flow,
            Average_Pressure,
            testing_duration,
            user_Id
        FROM Recipe_Log
        WHERE Timestamp >= ? AND Timestamp <= ?
        ORDER BY Timestamp DESC
    """

    cursor.execute(query, (start_date_str, end_date_str))
    rows = cursor.fetchall()

    recipe_logs = [
        {
            "SerialNo": row[0],
            "Timestamp": row[1],
            "Article_No": row[2],
            "Recipe_Name": row[3],
            "Filter_Code": row[4],
            "NgStatus": row[5],
            "RPM_Upper": row[6],
            "RPM_Lower": row[7],
            "Pressure_Upper": row[8],
            "Pressure_Lower": row[9],
            "AirFlow_Upper": row[10],
            "AirFlow_Lower": row[11],
            "Pressure_Max": row[12],
            "Pressure_Min": row[13],
            "Avg_Motor_RPM": row[14],
            "Avg_Air_Flow": row[15],
            "Avg_Pressure_Diff": row[16],
            "Test_Duration": row[17],
            "User_Id": row[18] if row[18] else "Admin",
        }
        for row in rows
    ]

    conn.close()
    return jsonify(recipe_logs)


@app.route("/export_recipe_log", methods=["GET"])
def export_recipe_log():
    conn = None
    try:
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed", "danger")
            return redirect(url_for("get_recipe_log"))

        cursor = conn.cursor()

        # Get date filters
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")

        # Convert to datetime objects for SQL Server
        if start_date:
            try:
                start_date = datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                start_date = None
        if end_date:
            try:
                end_date = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                end_date = None

        # Build query
        query = "SELECT * FROM Recipe_Log WHERE 1=1"
        params = []
        if start_date:
            query += " AND Timestamp >= ?"
            params.append(start_date)
        if end_date:
            query += " AND Timestamp <= ?"
            params.append(end_date)

        cursor.execute(query, params)
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description] if cursor.description else []
        conn.close()

        # Excel workbook (horizontal layout)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recipe Log"

        def normalize(val):
            return val if val not in ("", None) else None

        # Write header
        if columns:
            ws.append(columns)

        # Write data rows
        for row in rows:
            ws.append([normalize(val) for val in row])

        # Alignment: center horizontally and vertically
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value), default=0
            )
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="Recipe_Log.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        if conn:
            conn.close()
        flash(f"Error exporting Excel: {str(e)}", "danger")
        return redirect(url_for("get_recipe_log"))


# import datetime
# from opcua import Client, ua
BASE_NODE = 'ns=3;s="dbRecipe1"."Recipe"'
from datetime import datetime
import time


def update_batch_status():
    """Continuously reads OPC UA values and inserts data into recipe_log."""
    client = Client(ENDPOINT_URL)

    try:
        client.connect()
        print("Connected to OPC UA Server")

        while True:  # Continuous loop
            try:
                filter_complete_field_path = f'ns=3;s="dbReport"."filterComplete"'
                pack_number_field_path = f'ns=3;s="dbReport"."packNumber"'
                NODE_RECIPE_ID = f'ns=3;s="dbRecipe1"."Recipe"."RecipeID"'
                NODE_RECIPE_NAME = f'ns=3;s="dbRecipe1"."Recipe"."RecipeName"'

                # Fetch values from PLC
                # current_quantity = client.get_node(quantity_field_path).get_value()
                # machine_state = client.get_node(machine_state_field_path).get_value()
                # total_quantity = client.get_node(total_quantity_path).get_value()
                ###
                filter_complete = 1
                # pack_number=123;
                filter_complete = client.get_node(
                    filter_complete_field_path
                ).get_value()

                if filter_complete == True:  # When filterComplete is True

                    recipe_id = client.get_node(NODE_RECIPE_ID).get_value()
                    recipe_name = client.get_node(NODE_RECIPE_NAME).get_value()
                    pack_number = client.get_node(pack_number_field_path).get_value()
                    filter_complete_node = client.get_node(filter_complete_field_path)
                    filter_complete_node.set_value(
                        ua.DataValue(ua.Variant(False, ua.VariantType.Boolean))
                    )

                    # Generate Serial Number
                    today_date = datetime.now().strftime("%y%m%d")

                    serial_no = f"{today_date}{pack_number}"
                    batch_code = (
                        f"BATCH-{recipe_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    )
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    # printdata(recipe_name, article_no, serial_no)  # Call function

                    # Fetch additional data from `recipe` table
                    conn = get_db_connection()
                    if conn:
                        cursor = conn.cursor()
                        cursor.execute(
                            """
                    SELECT 
                    R.Recipe_ID, R.Recipe_Name, R.Filter_Size, R.Filter_Code, R.Art_No,
                     RD.Pos1, RD.Pos2, RD.Pos3, RD.Pos4, RD.Pos5, RD.Pos6, RD.Pos7, RD.Pos8, RD.Pos9,
                    RD.Alu_Material, RD.House_Material,
                     SM.Pleat_Height, SM.Soft_Touch, SM.Feeder1_Media_Thickness, SM.Feeder1_Park_Position,
                   SM.Foil1_Length, SM.Foil1_Length_Offset, SM.Foil1_Width, SM.Puller_Start_Position, SM.Autofeeding_Offset,
                 SM.Filter_Box_Height, SM.Filter_Box_Length, SM.Pleat_Pitch, SM.Pleat_Counts, SM.Foil_Low_Diameter,
                  SM.Feeding_Conveyor_Speed, SM.Pack_Transfer_Rev_Position, SM.Foil1_Tension_Set_Point, SM.Foil2_Tension_Set_Point,
                   SM.Blade_Opening, SM.Press_Touch, SM.Feeder2_Media_Thickness, SM.Feeder2_Park_Position,
                   SM.Foil2_Length, SM.Foil2_Length_Offset, SM.Foil2_Width, SM.Puller_End_Position,
                  SM.Puller2_Feed_Correction, SM.Puller_Extra_Stroke_Enable, SM.Filter_Box_Width,
                   SM.Lid_Placement_Enable, SM.Lid_Placement_Position, SM.Sync_Table_Start_Position,
                   SM.Batch_Count, SM.Discharge_Conveyor_Speed, SM.Pack_Transfer_Park_Position, SM.Media_Tension_Set_Point,
                    IS1.databaseAvailable, IS1.Width, IS1.Height, IS1.Depth, IS1.Art_No AS Inspection_Art_No,
                       IS1.Air_Flow_Set, IS1.Pressure_Drop_Setpoint, IS1.Lower_Tolerance1, IS1.Lower_Tolerance2,
                    IS1.Upper_Tolerance1, IS1.Upper_Tolerance2,IS1.Lower_Fan_Speed, IS1.Upper_Fan_Speed
                   FROM Recipe R
                      LEFT JOIN Recipe_Details1 RD ON R.Recipe_ID = RD.Recipe_ID
                            LEFT JOIN Sub_Menu SM ON R.Recipe_ID = SM.Recipe_ID
                         LEFT JOIN Inspection_Settings IS1 ON R.Recipe_ID = IS1.Recipe_ID
                           WHERE R.Recipe_ID = ?
                         """,
                            (recipe_id,),
                        )
                        recipe_full_data = cursor.fetchone()

                        # recipe_data = cursor.fetchone()

                        if recipe_full_data:

                            # Insert data into `recipe_log`
                            cursor.execute(
                                """
                        INSERT INTO Recipe_Log (
                        Batch_Code, Timestamp, SerialNo, NgStatus, Avg_Air_Flow, Avg_Result, Batch_Completion_Status,
                        Recipe_Id, Recipe_Name, Filter_Size, Filter_Code, Art_No,
                        Pos1, Pos2, Pos3, Pos4, Pos5, Pos6, Pos7, Pos8, Pos9, Alu_Material, House_Material,
                        Pleat_Height, Soft_Touch, Feeder1_Media_Thickness, Feeder1_Park_Position,
                        Foil1_Length, Foil1_Length_Offset, Foil1_Width, Puller_Start_Position, Autofeeding_Offset,
                        Filter_Box_Height, Filter_Box_Length, Pleat_Pitch, Pleat_Counts, Foil_Low_Diameter,
                        Feeding_Conveyor_Speed, Pack_Transfer_Rev_Position, Foil1_Tension_Set_Point, Foil2_Tension_Set_Point,
                        Blade_Opening, Press_Touch, Feeder2_Media_Thickness, Feeder2_Park_Position,
                        Foil2_Length, Foil2_Length_Offset, Foil2_Width, Puller_End_Position,
                        Puller2_Feed_Correction, Puller_Extra_Stroke_Enable, Filter_Box_Width,
                        Lid_Placement_Enable, Lid_Placement_Position, Sync_Table_Start_Position,
                        Batch_Count, Discharge_Conveyor_Speed, Pack_Transfer_Park_Position, Media_Tension_Set_Point,
                        DatabaseAvailable, Width, Height, Depth, Inspection_Art_No, Air_Flow_Set,
                        Pressure_Drop_Setpoint, Lower_Tolerance1, Lower_Tolerance2, Upper_Tolerance1, Upper_Tolerance2,Lower_Fan_Speed,Upper_Fan_Speed,user_Id
                        ) VALUES ( ?, GETDATE(),?,?,?,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?,?,?)""",
                                (
                                    batch_code,
                                    serial_no,
                                    "NA",
                                    "NA",
                                    "NA",
                                    "Completed",
                                    recipe_full_data.Recipe_ID,
                                    recipe_full_data.Recipe_Name,
                                    recipe_full_data.Filter_Size,
                                    recipe_full_data.Filter_Code,
                                    recipe_full_data.Art_No,
                                    recipe_full_data.Pos1,
                                    recipe_full_data.Pos2,
                                    recipe_full_data.Pos3,
                                    recipe_full_data.Pos4,
                                    recipe_full_data.Pos5,
                                    recipe_full_data.Pos6,
                                    recipe_full_data.Pos7,
                                    recipe_full_data.Pos8,
                                    recipe_full_data.Pos9,
                                    recipe_full_data.Alu_Material,
                                    recipe_full_data.House_Material,
                                    recipe_full_data.Pleat_Height,
                                    recipe_full_data.Soft_Touch,
                                    recipe_full_data.Feeder1_Media_Thickness,
                                    recipe_full_data.Feeder1_Park_Position,
                                    recipe_full_data.Foil1_Length,
                                    recipe_full_data.Foil1_Length_Offset,
                                    recipe_full_data.Foil1_Width,
                                    recipe_full_data.Puller_Start_Position,
                                    recipe_full_data.Autofeeding_Offset,
                                    recipe_full_data.Filter_Box_Height,
                                    recipe_full_data.Filter_Box_Length,
                                    recipe_full_data.Pleat_Pitch,
                                    recipe_full_data.Pleat_Counts,
                                    recipe_full_data.Foil_Low_Diameter,
                                    recipe_full_data.Feeding_Conveyor_Speed,
                                    recipe_full_data.Pack_Transfer_Rev_Position,
                                    recipe_full_data.Foil1_Tension_Set_Point,
                                    recipe_full_data.Foil2_Tension_Set_Point,
                                    recipe_full_data.Blade_Opening,
                                    recipe_full_data.Press_Touch,
                                    recipe_full_data.Feeder2_Media_Thickness,
                                    recipe_full_data.Feeder2_Park_Position,
                                    recipe_full_data.Foil2_Length,
                                    recipe_full_data.Foil2_Length_Offset,
                                    recipe_full_data.Foil2_Width,
                                    recipe_full_data.Puller_End_Position,
                                    recipe_full_data.Puller2_Feed_Correction,
                                    recipe_full_data.Puller_Extra_Stroke_Enable,
                                    recipe_full_data.Filter_Box_Width,
                                    recipe_full_data.Lid_Placement_Enable,
                                    recipe_full_data.Lid_Placement_Position,
                                    recipe_full_data.Sync_Table_Start_Position,
                                    recipe_full_data.Batch_Count,
                                    recipe_full_data.Discharge_Conveyor_Speed,
                                    recipe_full_data.Pack_Transfer_Park_Position,
                                    recipe_full_data.Media_Tension_Set_Point,
                                    recipe_full_data.databaseAvailable,
                                    recipe_full_data.Width,
                                    recipe_full_data.Height,
                                    recipe_full_data.Depth,
                                    recipe_full_data.Inspection_Art_No,
                                    recipe_full_data.Air_Flow_Set,
                                    recipe_full_data.Pressure_Drop_Setpoint,
                                    recipe_full_data.Lower_Tolerance1,
                                    recipe_full_data.Lower_Tolerance2,
                                    recipe_full_data.Upper_Tolerance1,
                                    recipe_full_data.Upper_Tolerance2,
                                    recipe_full_data.Lower_Fan_Speed,
                                    recipe_full_data.Upper_Fan_Speed,
                                    session.get("username", "Admin"),
                                ),
                            )

                            conn.commit()
                            print("Going for print")
                            printdata(recipe_name, recipe_full_data.Art_No, serial_no)
                            print(
                                f"Inserted into Recipe_Log: {recipe_id}, {recipe_name}, {serial_no}"
                            )

                        else:
                            print(f"Recipe ID {recipe_id} not found in database.")

                        cursor.close()
                        conn.close()

                    # Reset filterComplete node in OPC UA

                time.sleep(1)  # Wait 1 second before checking again

            except Exception as e:
                print(f"Error in OPC UA read: {e}")
                time.sleep(2)  # Retry after delay

    except Exception as e:
        print(f"OPC UA Connection Error: {e}")

    finally:
        client.disconnect()
        print("Disconnected from OPC UA Server")

Test_Complete = 'ns=3;s="dbReport"."airDropTestData".'
Test_Complete1 = 'ns=3;s="dbReport"."forAirDropTest".'
Test_Complete2 = 'ns=3;s="dbReport".'
Test_Complete3 = 'ns=3;s="dbCrossPlcComm"."airToMc".'

STATUS_MAP = {
    1: "FILTER OK",
    2: "FILTER OK",
    3: "FILTER OK C",
    4: "FILTER NG",
    5: "FILTER NG",
    8: "NG AIRFLOW LESS",
    9: "NG AIRFLOW GREATER",
}


def opcua_monitoring():
    """
    Continuously read OPC UA values, update MSSQL, and reset testComplete flag.
    """
    client = Client(ENDPOINT_URL)

    try:
        client.connect()
        print("Connected to OPC UA Server")

        while True:
            try:
                # 1) Read the testComplete flag
                test_complete_node = client.get_node(f"{Test_Complete}testComplete")
                testComplete = test_complete_node.get_value()

                if not testComplete:
                    time.sleep(1)
                    continue

                print("TestComplete is true")

                # 2) Read required OPC UA values
                serial_no = client.get_node(
                    f"{Test_Complete2}qrCodeDropAir_1"
                ).get_value()
                # avgAirFlow  = client.get_node(f"{Test_Complete}avgAirFlow").get_value()
                avgResult = client.get_node(f"{Test_Complete}avgResult").get_value()
                testResult = client.get_node(f"{Test_Complete}testResult").get_value()

                # 2a) Read extra values from testComplete3
                testing_duration = client.get_node(
                    f"{Test_Complete3}testing_Duration"
                ).get_value()
                avg_motor_rpm = client.get_node(
                    f"{Test_Complete3}Average_motor_RPM"
                ).get_value()
                airflow_lower = client.get_node(
                    f"{Test_Complete3}Airflow_Lower_limit"
                ).get_value()
                airflow_upper = client.get_node(
                    f"{Test_Complete3}Airflow_upper_limit"
                ).get_value()
                avg_airflow = client.get_node(
                    f"{Test_Complete3}Avg_Airflow"
                ).get_value()
                avg_pressure = client.get_node(
                    f"{Test_Complete3}avgPressure"
                ).get_value()
                lower_rpm = client.get_node(
                    f"{Test_Complete3}Upper_fan_Speed"
                ).get_value()
                upper_rpm = client.get_node(
                    f"{Test_Complete3}Lower_fan_speed"
                ).get_value()

                print("Serial:", serial_no)
                print("avgResult:", avgResult)
                print("testResult:", testResult)

                # 3) Determine NG status from mapping
                ng_status = STATUS_MAP.get(testResult, "NA")
                print("NG Status:", ng_status)

                # 4) Fetch Recipe_ID from database
                recipe_id = 0
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT Recipe_ID FROM recipe_log WHERE SerialNo = ?", (serial_no,)
                )
                row = cursor.fetchone()
                if row:
                    recipe_id = row[0]

                # 5) Reset the testComplete flag on OPC UA server
                test_complete_node.set_value(
                    ua.DataValue(ua.Variant(False, ua.VariantType.Boolean))
                )

                # 6) Update recipe log in DB
                update_recipe_log(
                    serial_no,
                    avgResult,
                    ng_status,
                    recipe_id,
                    testing_duration,
                    avg_motor_rpm,
                    airflow_lower,
                    airflow_upper,
                    avg_airflow,
                    avg_pressure,
                    lower_rpm,
                    upper_rpm,
                )

                conn.close()

            except Exception as e:
                print(f"Error reading/writing OPC UA: {e}")
                time.sleep(2)

            time.sleep(1)  # Polling delay

    except Exception as e:
        print(f"OPC UA Connection Error: {e}")

    finally:
        client.disconnect()
        print("Disconnected from OPC UA Server")


def opcua_qrcode_monitoring():
    """Continuously reads qrCodeDropAir, fetches Recipe ID, writes inspection settings, and resets qrCodeDropAir."""

    client = Client(ENDPOINT_URL)

    try:
        client.connect()
        print("✅ Connected to OPC UA Server (qrCode monitoring)")

        while True:
            try:
                # Step 1: Read qrCodeDropAir
                qr_code_node = client.get_node(f'{Test_Complete}"qrCodeDropAir"')

                qr_code_value = qr_code_node.get_value()

                if (
                    qr_code_value
                    and qr_code_value != "NO READ"
                    and qr_code_value != "NO READ                       "
                ):  # Not empty / not zero
                    print(f"📌 New QR Code detected: {qr_code_value}")
                    qr_code2_node = client.get_node(
                        f'{Test_Complete2}"qrCodeDropAir_1"'
                    )
                    qr_code2_node.set_value(
                        ua.DataValue(ua.Variant(qr_code_value, ua.VariantType.String))
                    )
                    print(f"➡️ QR Code copied to qrCode2: {qr_code_value}")
                    # Step 2: Fetch Recipe_ID using qr_code_value
                    recipe_id = 0
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    cursor.execute(
                        "SELECT Recipe_ID FROM recipe_log WHERE SerialNo = ?",
                        (qr_code_value,),
                    )
                    row = cursor.fetchone()

                    if row:
                        recipe_id = row[0]

                    if recipe_id:
                        # Step 3: Fetch inspection settings for the Recipe_ID
                        cursor.execute(
                            """
                            SELECT databaseAvailable, Width, Height, Depth, Art_No, Air_Flow_Set, 
                                   Pressure_Drop_Setpoint, Lower_Tolerance1, Lower_Tolerance2, 
                                   Upper_Tolerance1, Upper_Tolerance2
                            FROM Recipe_Log WHERE Recipe_ID = ?
                        """,
                            (recipe_id,),
                        )
                        settings = cursor.fetchone()

                        if settings:
                            # Step 4: Write inspection settings to PLC
                            opcua_nodes = [
                                "databaseAvailable",
                                "Width",
                                "Height",
                                "Depth",
                                "Art No.",
                                "Air Flow set",
                                "Pressure Drop Setpoint",
                                "Lower_Tolerance1",
                                "Lower_Tolerance2",
                                "Upper_Tolerance1",
                                "Upper_Tolerance2",
                            ]

                            for i, submenu_value in enumerate(settings):
                                node = client.get_node(
                                    f'{Test_Complete1}"{opcua_nodes[i]}"'
                                )
                                data_type = node.get_data_type_as_variant_type()

                                if data_type == ua.VariantType.Float:
                                    value = float(submenu_value)
                                elif data_type in (
                                    ua.VariantType.Int32,
                                    ua.VariantType.Int16,
                                ):
                                    value = int(submenu_value)
                                elif data_type == ua.VariantType.Boolean:
                                    value = bool(submenu_value)
                                else:
                                    print(
                                        f"⚠️ Skipping {opcua_nodes[i]} due to unknown data type: {data_type}"
                                    )
                                    continue

                                node.set_value(
                                    ua.DataValue(ua.Variant(value, data_type))
                                )

                            print(
                                f"✅ Inspection settings written for Recipe_ID {recipe_id}"
                            )
                        else:
                            print(
                                f"⚠️ No Inspection Settings found for Recipe_ID {recipe_id}"
                            )

                    conn.close()

                    # Step 5: Reset qrCodeDropAir → 0
                    qr_code_node.set_value(
                        ua.DataValue(ua.Variant("", ua.VariantType.String))
                    )
                    print("🔄 qrCodeDropAir reset to 0")

                time.sleep(1)

            except Exception as e:
                print(f"❌ Error in qrCode monitoring: {e}")
                time.sleep(2)

    except Exception as e:
        print(f"❌ OPC UA Connection Error: {e}")

    finally:
        client.disconnect()
        print("🔌 Disconnected from OPC UA Server (qrCode monitoring)")

#need to update this funtion when i will get info about which fileds are available and at which plc location 
def update_recipe_log(
    serial_no,
    avgResult,
    ng_status,
    recipe_id,
    testing_duration,
    avg_motor_rpm,
    airflow_lower,
    airflow_upper,
    avg_airflow,
    avg_pressure,
    lower_rpm,
    upper_rpm,
):
    """Updates the MSSQL recipe_log table with additional PLC values."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute(
            "SELECT recipe_id FROM Recipe_Log WHERE SerialNo = ?", (serial_no,)
        )
        row = cursor.fetchone()

        if row:
            # Update existing record
            cursor.execute(
                """
                UPDATE Recipe_Log
                SET Avg_Air_Flow = ?, Avg_Result = ?, NgStatus = ?,
                    testing_duration = ?, Average_motor_RPM = ?, 
                    Airflow_lower_limit = ?, Airflow_upper_limit = ?, 
                    Average_Pressure = ?, Lower_fan_Speed = ?, Upper_fan_Speed = ?
                WHERE SerialNo = ?
            """,
                (
                    avg_airflow,
                    avgResult,
                    ng_status,
                    testing_duration,
                    avg_motor_rpm,
                    airflow_lower,
                    airflow_upper,
                    avg_pressure,
                    lower_rpm,
                    upper_rpm,
                    serial_no,
                ),
            )
            conn.commit()
            print(f"Updated record for Serial No: {serial_no}")

        else:
            # Insert new record
            batch_code = f"BATCH-{recipe_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}"

            cursor.execute(
                """
                INSERT INTO Recipe_Log 
                (Batch_Code, Timestamp, Recipe_Id, Recipe_Name, Arte_No, 
                 Filter_Size, FilterSize, NgStatus, SerialNo, 
                 Avg_Air_Flow, Avg_Result, Batch_Completion_Status,
                 testing_duration, Average_motor_RPM, Airflow_lower_limit, 
                 Airflow_upper_limit, Average_Pressure,Lower_fan_Speed,Upper_fan_Speed)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?,?)
            """,
                (
                    batch_code,
                    datetime.now(),
                    recipe_id,
                    "NA",
                    "NA",
                    "NA",
                    "NA",
                    ng_status,
                    serial_no,
                    avg_airflow,
                    avgResult,
                    "Completed",
                    testing_duration,
                    avg_motor_rpm,
                    airflow_lower,
                    airflow_upper,
                    avg_pressure,
                    lower_rpm,
                    upper_rpm,
                ),
            )

            conn.commit()
            print(f"Inserted new record for Serial No: {serial_no}")

        cursor.close()
        conn.close()

    except Exception as e:
        print(f"Database error: {e}")


# Printer Configuration
PRINTER_HOST = "192.168.0.61"
PRINTER_PORT = 6101


@app.route("/api/print_recipe/<serial_no>", methods=["POST"])
def print_recipe(serial_no):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Fetch RecipeName and ArticleNo from Recipe_Log based on SerialNo
        cursor.execute(
            """
            SELECT Recipe_Name, Art_No 
            FROM Recipe_Log
            WHERE SerialNo = ?
        """,
            (serial_no,),
        )
        recipe_data = cursor.fetchone()

        if recipe_data:
            recipe_name, article_number = recipe_data
            printdata(
                recipe_name, article_number, serial_no
            )  # Call your print function

            # Respond with success message
            return jsonify(
                {"success": True, "message": "Printing initiated successfully."}
            )
        else:
            return jsonify(
                {
                    "success": False,
                    "message": "Serial number not found in the recipe log.",
                }
            )

    except Exception as e:
        print(f"Error: {e}")
        return jsonify(
            {"success": False, "message": "Error occurred while processing print."}
        )
    finally:
        cursor.close()
        conn.close()


def send_to_printer(data):
    """Function to send data to the printer via a TCP socket."""
    client = None
    try:
        client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        client.settimeout(5)  # Prevent hanging
        client.connect((PRINTER_HOST, PRINTER_PORT))

        print("Sending data to printer...")  # Debugging

        client.sendall(data.encode("utf-8"))  # Send data

        try:
            response = client.recv(1024).decode("utf-8")  # Get response
            print(f"Response from printer: {response}")
        except socket.timeout:
            print("Printer did not send a response (this is normal for some printers).")

    except Exception as e:
        print(f"Error sending data to printer: {e}")
    finally:
        if client:
            client.close()
        print("Printer connection closed.")


def printdata(recipe_name, article_no, serial_no):
    """Function to print the slip and generate a PDF."""
    print_data = f"""  
^XA
~TA000
~JSN
^LT0
^MNW
^MTD
^PON
^PMN
^LH0,0
^JMA
^PR4,4
~SD15
^JUS
^LRN
^CI27
^PA0,1,1,0
^XZ
^XA
^MMT
^PW900
^LL600
^LS0
^FT301,54^A0I,103,91^FH\\^CI28^FDV^FS^CI27
^FT301,143^A0N,85,86^FH\\^CI28^FDbsolent^FS^CI27
^FT360,82^A0N,76,109^FH\\^CI28^FD>^FS^CI27
^FT606,116^A0N,21,20^FH\\^CI28^FDpart of Abosent^FS^CI27
^FT606,142^A0N,21,20^FH\\^CI28^FDAir Care Group^FS^CI27
^FT13,247^A0N,44,53^FH\\^CI28^FDType :^FS^CI27
^FT13,314^A0N,49,38^FH\\^CI28^FDAbsolent filter ^FS^CI27
^FT280,312^A0N,54,43^FH\\^CI28^FD{recipe_name}^FS^CI27
^FT13,372^A0N,39,56^FH\\^CI28^FDArt.no : ^FS^CI27
^FT13,461^A0N,67,74^FH\\^CI28^FD{article_no}^FS^CI27
^FT13,556^A0N,34,38^FH\\^CI28^FDSerial no :^FS^CI27
^FT196,567^A0N,53,20^FH\\^CI28^FD{serial_no}^FS^CI27
^FT537,563^BQN,2,8
^FH\\^FDLA,{serial_no}^FS
^FT859,557^A0B,65,109^FH\\^CI28^FDAirflow^FS^CI27
^FO838,175^GB0,58,4^FS
^FT824,214^A0N,71,71^FH\\^CI28^FD\\5E^FS^CI27
^PQ1,0,0,N
^XZ`
    """

    # Send the formatted data to the printer
    send_to_printer(print_data)


@app.route("/role", methods=["GET", "POST"])
def role():
    """Manage user roles (Admin only)."""
    if session.get("username") != "Admin":  # Ensure only Admins can access
        flash("Access denied: Admins only.", "danger")
        return redirect(url_for("dashboard"))

    conn = get_db_connection()
    if not conn:
        return redirect(url_for("dashboard"))

    cursor = conn.cursor()

    # Fetch unique usernames
    cursor.execute("SELECT DISTINCT username FROM users")
    users = [row[0] for row in cursor.fetchall()]

    if request.method == "POST":
        selected_user = request.form.get("username")
        roles = request.form.getlist("roles")  # Get multiple roles as a list

        if selected_user:
            roles_str = ",".join(roles)  # Convert list to comma-separated string

            # Update roles in the database
            cursor.execute(
                """
                UPDATE users SET roles = ? WHERE username = ?
            """,
                (roles_str, selected_user),
            )

            conn.commit()
            flash(f"Roles updated for {selected_user}", "success")

    cursor.close()
    conn.close()
    return render_template("roles.html", users=users)


@app.route("/get_roles", methods=["POST"])
def get_roles():
    """Fetch user roles from MSSQL based on the provided username."""
    username = request.json.get("username")  # Get username from the AJAX request
    if not username:
        return jsonify({"error": "Username is required"}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500

    cursor = conn.cursor()

    try:
        # Fetch roles for the selected user
        cursor.execute("SELECT roles FROM users WHERE username = ?", (username,))
        result = cursor.fetchone()

        roles = result[0].split(",") if result and result[0] else []
        return jsonify({"roles": roles})

    except pyodbc.Error as e:
        return jsonify({"error": f"Database query error: {str(e)}"}), 500

    finally:
        cursor.close()
        conn.close()


@app.route("/event_log", methods=["GET"])
def event_log():
    """Fetch and display user login/logout activity logs with pagination."""
    if session.get("username") != "Admin":  # Replace 'Admin' with your admin identifier
        flash("Access denied: Admins only.", "danger")
        return redirect(url_for("dashboard"))

    conn = get_db_connection()
    if not conn:
        return redirect(url_for("dashboard"))

    cursor = conn.cursor()

    # Fetch user summary data for activity logs (handling logout properly)
    cursor.execute(
        """
        SELECT 
            username,
            MAX(login_time) AS last_login_time,
            (
                SELECT TOP 1 logout_time 
                FROM user_activity 
                WHERE username = ua.username 
                AND logout_time IS NOT NULL 
                ORDER BY logout_time DESC
            ) AS last_logout_time
        FROM user_activity ua
        GROUP BY username
    """
    )

    user_summary = []
    for row in cursor.fetchall():
        username, last_login_time, last_logout_time = row
        last_login_date = (
            last_login_time.strftime("%Y-%m-%d") if last_login_time else "N/A"
        )
        last_login_time = (
            last_login_time.strftime("%H:%M:%S") if last_login_time else "N/A"
        )
        last_logout_date = (
            last_logout_time.strftime("%Y-%m-%d") if last_logout_time else "N/A"
        )
        last_logout_time = (
            last_logout_time.strftime("%H:%M:%S") if last_logout_time else "N/A"
        )

        user_summary.append(
            {
                "username": username,
                "last_login_date": last_login_date,
                "last_login_time": last_login_time,
                "last_logout_date": last_logout_date,
                "last_logout_time": (
                    last_logout_time if last_logout_time else "Currently Logged In"
                ),
            }
        )

    # Pagination setup
    page = request.args.get("page", 1, type=int)
    per_page = 5
    offset = (page - 1) * per_page

    selected_username = request.args.get("username")
    selected_user_details = []
    total_pages = None

    if selected_username:
        cursor.execute(
            """
            SELECT 
                CONVERT(VARCHAR, login_time, 23) AS login_date,
                CONVERT(VARCHAR, login_time, 8) AS login_time,
                CASE 
                    WHEN logout_time IS NULL THEN NULL 
                    ELSE CONVERT(VARCHAR, logout_time, 23) 
                END AS logout_date,
                CASE 
                    WHEN logout_time IS NULL THEN NULL 
                    ELSE CONVERT(VARCHAR, logout_time, 8) 
                END AS logout_time
            FROM user_activity
            WHERE username = ?
            ORDER BY login_time DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """,
            (selected_username, offset, per_page),
        )

        selected_user_details = [
            {
                "login_date": row[0],
                "login_time": row[1],
                "logout_date": row[2] or "Currently Logged In",
                "logout_time": row[3] or "Currently Logged In",
            }
            for row in cursor.fetchall()
        ]

        # Get total log count for pagination
        cursor.execute(
            "SELECT COUNT(*) FROM user_activity WHERE username = ?",
            (selected_username,),
        )
        total_logs = cursor.fetchone()[0]
        total_pages = (total_logs + per_page - 1) // per_page  # Calculate total pages

    conn.close()

    return render_template(
        "event_log.html",
        user_summary=user_summary,
        selected_username=selected_username,
        selected_user_details=selected_user_details,
        page=page,
        total_pages=total_pages,
    )


@app.route("/reset-password", methods=["POST"])
def reset_password():

    user_id = request.form["user_id"]
    new_password = request.form["new_password"]

    # Hash the new password (important for security)
    hashed_password = generate_password_hash(new_password)

    # Update the password in the database
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET password = ? WHERE id = ?", (hashed_password, user_id)
    )
    conn.commit()
    conn.close()

    flash("Password has been reset successfully!", "success")
    return redirect(url_for("users"))


@app.route("/login", methods=["GET", "POST"])
def login():
    """User login with password hashing, role-based access, and activity tracking."""
    conn = get_db_connection()
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        if not conn:
            return redirect(url_for("login"))

        cursor = conn.cursor()

        # Fetch user details along with roles
        cursor.execute(
            "SELECT id, username, password, is_admin, roles FROM users WHERE username = ?",
            (username,),
        )
        user = cursor.fetchone()

        if user and check_password_hash(user[2], password):  # Verifying hashed password
            session["id"] = user[0]
            session["username"] = user[1]
            session["is_admin"] = user[3]

            # Store roles in session as a list (split by comma)
            session["user_roles"] = user[4].split(",") if user[4] else []

            # Insert login activity (MSSQL uses `GETDATE()` for current timestamp)
            cursor.execute(
                "INSERT INTO user_activity (username, login_time) VALUES (?, GETDATE())",
                (username,),
            )
            conn.commit()

            flash("Login successful!", "success")
            return redirect(url_for("dashboard"))

        else:
            flash("Invalid username or password.", "danger")

        conn.close()

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    """User registration with hashed passwords and duplicate username/email check."""
    if request.method == "POST":
        username = request.form.get("username")
        email = request.form.get("email")
        password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")

        # Check if passwords match
        if password != confirm_password:
            flash("Passwords do not match!", "danger")
            return redirect(url_for("register"))

        # Hash the password
        hashed_password = generate_password_hash(password)

        conn = get_db_connection()
        if not conn:
            return redirect(url_for("register"))

        try:
            cursor = conn.cursor()

            # Check if username or email already exists
            cursor.execute(
                "SELECT 1 FROM users WHERE username = ? OR email = ?", (username, email)
            )
            if cursor.fetchone():
                flash("Username or email already taken!", "danger")
                conn.close()
                return redirect(url_for("register"))

            # Insert new user into MSSQL database
            cursor.execute(
                "INSERT INTO users (username, email, password) VALUES (?, ?, ?)",
                (username, email, hashed_password),
            )
            conn.commit()

            flash("Registration successful! Please log in.", "success")
            return redirect(url_for("login"))

        except pyodbc.Error as e:
            flash(f"Database error: {e}", "danger")
        finally:
            print("REGISTER DONE")

    return render_template("register.html")


@app.route("/users")
def users():
    """Admin-only route to fetch user details from MSSQL database."""
    if session.get("username") != "Admin":  # Ensure only Admin can access
        flash("Access denied: Admins only.", "danger")
        return redirect(url_for("dashboard"))

    conn = get_db_connection()
    if not conn:
        return redirect(url_for("dashboard"))

    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id, username, email, is_admin FROM users")
        users_data = cursor.fetchall()

        users_list = []
        for row in users_data:
            users_list.append(
                {
                    "id": row[0],
                    "username": row[1],
                    "email": row[2],
                    "is_admin": "Yes" if row[3] else "No",
                }
            )

        return render_template("users.html", users=users_list)

    except pyodbc.Error as e:
        flash(f"Database error: {e}", "danger")
    finally:
        conn.close()


@app.route("/validate-admin-password", methods=["POST"])
def validate_admin_password():
    data = request.json
    admin_password = "admin123"  # Predefined admin password

    if data.get("password") == admin_password:
        return jsonify({"success": True})
    else:
        return jsonify({"success": False})


@app.route("/logout")
def logout():
    """Logs out the user and updates their logout time in MSSQL."""
    if "username" in session:
        username = session["username"]

        conn = get_db_connection()
        if not conn:
            return redirect(url_for("dashboard"))  # Redirect if connection fails

        try:
            cursor = conn.cursor()

            # Update the logout time for the latest login session
            cursor.execute(
                """
                UPDATE user_activity 
                SET logout_time = GETDATE() 
                WHERE username = ? 
                AND logout_time IS NULL
            """,
                (username,),
            )
            conn.commit()

            # Remove user from session
            session.pop("username", None)

            flash("You have been logged out successfully.", "success")

        except pyodbc.Error as e:
            flash(f"Database error: {e}", "danger")
        finally:
            conn.close()

    return redirect(url_for("login"))


@app.route("/tag-overview")
def tag_overview():
    """Displays a paginated tag overview."""
    if "username" not in session:  # Check if the user is logged in
        flash("Login first to access the Tag Overview.", "warning")
        return redirect(url_for("login"))  # Redirect to login if not logged in

    # Get the page and limit parameters from the query string
    current_page = int(request.args.get("page", 1))  # Default to page 1
    limit = int(request.args.get("limit", 10))  # Default limit is 10

    # Calculate offset for pagination
    offset = (current_page - 1) * limit

    # Establish database connection
    conn = get_db_connection()
    if not conn:
        flash("Database connection failed.", "danger")
        return redirect(url_for("dashboard"))

    try:
        cursor = conn.cursor()

        # Fetch paginated data
        cursor.execute(
            """
            SELECT * FROM Tag_Table 
            ORDER BY id 
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """,
            (offset, limit),
        )
        tags = cursor.fetchall()

        # Fetch the total number of rows
        cursor.execute("SELECT COUNT(*) FROM Tag_Table")
        total_rows = cursor.fetchone()[0]
        total_pages = (total_rows + limit - 1) // limit  # Calculate total pages

    except pyodbc.Error as e:
        flash(f"Database error: {e}", "danger")
        return redirect(url_for("dashboard"))

    finally:
        conn.close()

    # Render the template with tags and pagination details
    return render_template(
        "tag_overview.html",
        tags=tags,
        page=current_page,
        total_pages=total_pages,
        limit=limit,
        active_menu="tags",
        active_submenu="tag-overview",
    )


@app.route("/update-tag/<string:tagId>", methods=["PUT"])
def update_tag(tagId):
    """Update an existing tag in the Tag_Table."""
    try:
        # Parse JSON data from the request
        data = request.get_json()
        tagName = data.get("tagName")
        tagAddress = data.get("tagAddress")
        plcId = data.get("plcId")

        if not tagName or not tagAddress or not plcId:
            return jsonify({"success": False, "error": "Missing required fields"}), 400

        # Connect to MSSQL database
        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "error": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Check if the tag exists
        cursor.execute("SELECT 1 FROM Tag_Table WHERE tagId = ?", (tagId,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({"success": False, "error": "Tag not found"}), 404

        # Update the tag details in the database
        cursor.execute(
            """
            UPDATE Tag_Table
            SET tagName = ?, tagAddress = ?, plcId = ?
            WHERE tagId = ?
        """,
            (tagName, tagAddress, plcId, tagId),
        )

        conn.commit()
        conn.close()

        # Fetch and update the live value of the updated tag
        fetch_and_update_live_value(tagAddress)

        # Respond with success message
        return jsonify({"success": True, "message": "Tag updated successfully"})

    except pyodbc.Error as e:
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500

    except Exception as e:
        return jsonify({"success": False, "error": f"Server error: {str(e)}"}), 500


@app.route("/delete-tag/<int:tag_id>", methods=["DELETE"])
def delete_tag(tag_id):
    """Delete a tag from Tag_Table and Live_Tags in MSSQL."""
    try:
        # Check if user is logged in
        if "username" not in session:
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        # Connect to MSSQL database
        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "error": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Execute deletion queries
        cursor.execute("DELETE FROM Tag_Table WHERE tagId = ?", (tag_id,))
        cursor.execute("DELETE FROM Live_Tags WHERE tagId = ?", (tag_id,))
        conn.commit()

        # Check if any rows were affected
        if cursor.rowcount == 0:
            return jsonify({"success": False, "error": "Tag not found"}), 404

        return jsonify({"success": True, "message": "Tag deleted successfully"})

    except pyodbc.Error as e:
        print(f"Database error: {e}")
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500

    except Exception as e:
        print(f"Error while deleting tag: {e}")
        return jsonify({"success": False, "error": f"Server error: {str(e)}"}), 500

    finally:
        if conn:
            conn.close()


@app.route("/live-tag")
def live_tag():
    """Fetch paginated live tag data from MSSQL."""
    if "username" not in session:  # Check if the user is logged in
        return redirect(url_for("login"))

    # Get the current page from the query string (default to 1)
    current_page = int(request.args.get("page", 1))

    # Pagination setup: 10 items per page
    items_per_page = 10
    offset = (current_page - 1) * items_per_page

    # Connect to MSSQL database
    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Database connection failed"}), 500

    cursor = conn.cursor()

    try:
        # Fetch paginated live tag data
        cursor.execute(
            """
            SELECT Live_Tags.Id, Tag_Table.tagName, Live_Tags.value, Live_Tags.timestamp
            FROM Live_Tags
            JOIN Tag_Table ON Live_Tags.tagId = Tag_Table.tagId
            ORDER BY Live_Tags.timestamp DESC
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """,
            (offset, items_per_page),
        )
        live_tags = cursor.fetchall()

        # Fetch total number of rows for pagination
        cursor.execute("SELECT COUNT(*) FROM Live_Tags")
        total_rows = cursor.fetchone()[0]
        total_pages = (
            total_rows + items_per_page - 1
        ) // items_per_page  # Calculate total pages

    except pyodbc.Error as e:
        print(f"Database error: {e}")
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500

    finally:
        conn.close()

    return render_template(
        "live_tag.html",
        live_tags=live_tags,  # Pass the retrieved live tag data
        page=current_page,  # Pass the current page number
        total_pages=total_pages,  # Pass total number of pages
        active_menu="tags",  # Keep the Tags dropdown open
        active_submenu="live-tag",  # Highlight the Live Tags submenu
    )


@app.route("/plc")
def plc():
    """Fetch PLC and Machine data from MSSQL."""
    if session.get("username") != "Admin":  # Ensure only Admin can access
        flash("Access denied: Admins only.", "danger")
        return redirect(url_for("dashboard"))

    # Connect to MSSQL database
    conn = get_db_connection()
    if not conn:
        flash("Database connection failed", "danger")
        return redirect(url_for("dashboard"))

    cursor = conn.cursor()

    try:
        # Fetch PLC data
        cursor.execute("SELECT * FROM Plc_Table")  # Adjust query as per your DB schema
        plc_tags = cursor.fetchall()

        # # Fetch Machine data
        # cursor.execute("SELECT * FROM Machine_Table")  # Adjust query as per your DB schema
        # machine_tags = cursor.fetchall()

        machine_tags = [
            {"machineId": "1", "machineName": "Machine1", "URL": "192.168.1.2"}
        ]

    except pyodbc.Error as e:
        print(f"Database error: {e}")
        flash(f"Database error: {str(e)}", "danger")
        return redirect(url_for("dashboard"))

    finally:
        conn.close()

    # Render the template with the PLC and Machine data
    return render_template("plc.html", plc_tags=plc_tags, machine_tags=machine_tags)


@app.route("/raw-material", methods=["GET", "POST"])
def raw_material():
    """Handle raw material addition (POST) and retrieval with pagination (GET)."""

    if "username" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        try:
            # Extract required fields from form
            type1 = request.form["type1"]
            width = request.form["width"]
            part_no = request.form["part_no"]

            # Auto-generate fields
            make = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            user = session.get("username")
            # barcode = f"-{type_code}-{lot_no}"  # Example barcode
            material_type = f"{type1}x{width}:{part_no}"
            barcode = material_type
            if not type1 or not width or not part_no:
                return jsonify({"error": "All fields are required"}), 400

            # Connect to MSSQL database
            conn = get_db_connection()
            if not conn:
                return (
                    jsonify(
                        {"success": False, "message": "Database connection failed"}
                    ),
                    500,
                )

            cursor = conn.cursor()

            # Insert data into the Raw_Materials table
            cursor.execute(
                """
                INSERT INTO Raw_Materials (type, width,part_no, materialType, make, [user], barcode) 
                VALUES (?, ?, ?, ?, ?, ?,?)""",
                (type1, width, part_no, material_type, make, user, barcode),
            )
            conn.commit()

            return jsonify({"success": True, "message": "Material added successfully!"})

        except pyodbc.Error as e:
            print(f"Database Error: {e}")
            return (
                jsonify({"success": False, "message": f"Database error: {str(e)}"}),
                500,
            )

        finally:
            conn.close()

    # Handle GET request for pagination
    try:
        current_page = int(request.args.get("page", 1))
        limit = int(request.args.get("limit", 10))
        offset = (current_page - 1) * limit

        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "message": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Fetch paginated raw materials
        cursor.execute(
            """
            SELECT * FROM Raw_Materials 
            ORDER BY material_Id
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """,
            (offset, limit),
        )
        raw_materials = cursor.fetchall()

        # Calculate total pages
        cursor.execute("SELECT COUNT(*) FROM Raw_Materials")
        total_rows = cursor.fetchone()[0]
        total_pages = (total_rows + limit - 1) // limit

        conn.close()

        # Render the template with paginated data
        return render_template(
            "raw_material.html",
            raw_materials=raw_materials,
            page=current_page,
            total_pages=total_pages,
            limit=limit,
        )

    except pyodbc.Error as e:
        print(f"Database Error: {e}")
        return jsonify({"success": False, "message": f"Database error: {str(e)}"}), 500


@app.route("/delete-raw-material/<int:raw_material_id>", methods=["DELETE"])
def delete_raw_material(raw_material_id):
    """Delete a raw material entry from MSSQL database."""

    try:
        print(f"Attempting to delete raw material with ID: {raw_material_id}")

        # Check if user is logged in
        if "username" not in session:
            print("Unauthorized access attempt.")
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        # Connect to the MSSQL database
        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "error": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Check if the material exists
        cursor.execute(
            "SELECT * FROM Raw_Materials WHERE material_Id = ?", (raw_material_id,)
        )
        row = cursor.fetchone()

        if row is None:
            print(f"Material with ID {raw_material_id} not found.")
            return jsonify({"success": False, "error": "Raw material not found"}), 404

        # Delete the raw material
        cursor.execute(
            "DELETE FROM Raw_Materials WHERE material_Id = ?", (raw_material_id,)
        )
        conn.commit()

        if cursor.rowcount == 0:
            print(f"Failed to delete material with ID {raw_material_id}.")
            return (
                jsonify({"success": False, "error": "Failed to delete material"}),
                500,
            )

        print(f"Successfully deleted material with ID {raw_material_id}.")
        return jsonify({"success": True})

    except pyodbc.Error as e:
        print(f"Database Error: {e}")
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500

    finally:
        if conn:
            conn.close()


@app.route("/update-raw-material/<int:material_Id>", methods=["POST"])
def update_raw_material(material_Id):
    """Update raw material details in MSSQL database."""
    conn = None  # Ensure conn is defined before try

    try:
        data = request.json
        print(data)

        # Validate required fields
        required_fields = ["type", "width", "part_no", "make", "user", "barcode"]
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Missing fields: {', '.join(missing_fields)}",
                    }
                ),
                400,
            )

        type1 = data.get("type")
        width = data.get("width")
        part_no = data.get("part_no")
        make = data.get("make")
        user = session.get("username")
        barcode = f"{type1}x{width}:{part_no}"
        materialType = f"{type1}x{width}:{part_no}"

        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "error": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Check if material exists
        cursor.execute(
            "SELECT COUNT(*) FROM Raw_Materials WHERE material_Id = ?", (material_Id,)
        )
        if cursor.fetchone()[0] == 0:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "Material with the given ID does not exist.",
                    }
                ),
                404,
            )

        # Perform update
        cursor.execute(
            """
            UPDATE Raw_Materials
            SET type = ?, width = ?, part_no = ?, make = ?, [user] = ?, materialType = ?, barcode = ?
            WHERE material_Id = ?
        """,
            (type1, width, part_no, make, user, materialType, barcode, material_Id),
        )
        conn.commit()

        if cursor.rowcount == 0:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "No changes were made to the raw material.",
                    }
                ),
                400,
            )

        return jsonify(
            {"success": True, "message": "Raw material updated successfully."}
        )

    except pyodbc.IntegrityError as e:
        if "UNIQUE constraint failed: Raw_Materials.materialType" in str(e):
            return (
                jsonify({"success": False, "error": "Material Type must be unique."}),
                409,
            )
        return (
            jsonify({"success": False, "error": f"Database integrity error: {str(e)}"}),
            400,
        )

    except pyodbc.Error as e:
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500

    except Exception as e:
        return (
            jsonify(
                {"success": False, "error": f"An unexpected error occurred: {str(e)}"}
            ),
            500,
        )

    finally:
        if conn:
            conn.close()


@app.route("/alu-raw-material", methods=["GET", "POST"])
def raw_material1():
    """Handle raw material addition (POST) and retrieval with pagination (GET)."""

    if "username" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        try:
            # Extract required fields from form
            type1 = request.form["thickness"]
            width = request.form["width"]
            part_no = request.form["part_no"]
            if not type1 or not width or not part_no:
                return jsonify({"error": "All fields are required"}), 400

            # Auto-generate fields
            make = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            user = session.get("username")
            # barcode = f"-{type_code}-{lot_no}"  # Example barcode
            material_type = f"{width}x{type1}:{part_no}"
            barcode = material_type

            # Connect to MSSQL database
            conn = get_db_connection()
            if not conn:
                return (
                    jsonify(
                        {"success": False, "message": "Database connection failed"}
                    ),
                    500,
                )

            cursor = conn.cursor()

            # Insert data into the Raw_Materials table
            cursor.execute(
                """
                INSERT INTO Alu_Raw_Materials (thickness, width,part_no, materialType, make, [user], barcode) 
                VALUES (?, ?, ?, ?, ?, ?,?)""",
                (type1, width, part_no, material_type, make, user, barcode),
            )
            conn.commit()

            return jsonify({"success": True, "message": "Material added successfully!"})

        except pyodbc.Error as e:
            print(f"Database Error: {e}")
            return (
                jsonify({"success": False, "message": f"Database error: {str(e)}"}),
                500,
            )

        finally:
            conn.close()

    # Handle GET request for pagination
    try:
        current_page = int(request.args.get("page", 1))
        limit = int(request.args.get("limit", 10))
        offset = (current_page - 1) * limit

        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "message": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Fetch paginated raw materials
        cursor.execute(
            """
            SELECT * FROM Alu_Raw_Materials 
            ORDER BY material_Id
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """,
            (offset, limit),
        )
        raw_materials = cursor.fetchall()

        # Calculate total pages
        cursor.execute("SELECT COUNT(*) FROM Alu_Raw_Materials")
        total_rows = cursor.fetchone()[0]
        total_pages = (total_rows + limit - 1) // limit

        conn.close()

        # Render the template with paginated data
        return render_template(
            "aluminium_raw_material.html",
            raw_materials=raw_materials,
            page=current_page,
            total_pages=total_pages,
            limit=limit,
        )

    except pyodbc.Error as e:
        print(f"Database Error: {e}")
        return jsonify({"success": False, "message": f"Database error: {str(e)}"}), 500


@app.route("/delete-alu_raw-material/<int:raw_material_id>", methods=["DELETE"])
def delete_raw_material1(raw_material_id):
    """Delete a raw material entry from MSSQL database."""

    try:
        print(f"Attempting to delete alu_raw material with ID: {raw_material_id}")

        # Check if user is logged in
        if "username" not in session:
            print("Unauthorized access attempt.")
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        # Connect to the MSSQL database
        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "error": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Check if the material exists
        cursor.execute(
            "SELECT * FROM Alu_Raw_Materials WHERE material_Id = ?", (raw_material_id,)
        )
        row = cursor.fetchone()

        if row is None:
            print(f"Material with ID {raw_material_id} not found.")
            return jsonify({"success": False, "error": "Raw material not found"}), 404

        # Delete the raw material
        cursor.execute(
            "DELETE FROM Alu_Raw_Materials WHERE material_Id = ?", (raw_material_id,)
        )
        conn.commit()

        if cursor.rowcount == 0:
            print(f"Failed to delete material with ID {raw_material_id}.")
            return (
                jsonify({"success": False, "error": "Failed to delete material"}),
                500,
            )

        print(f"Successfully deleted material with ID {raw_material_id}.")
        return jsonify({"success": True})

    except pyodbc.Error as e:
        print(f"Database Error: {e}")
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500

    finally:
        if conn:
            conn.close()


@app.route("/update-alu_raw-material/<int:material_Id>", methods=["POST"])
def update_raw_material1(material_Id):
    """Update raw material details in MSSQL database."""
    conn = None  # Ensure conn is defined before try

    try:
        data = request.json
        print(data)

        # Validate required fields
        required_fields = ["thickness", "width", "part_no", "make", "user", "barcode"]
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Missing fields: {', '.join(missing_fields)}",
                    }
                ),
                400,
            )

        type1 = data.get("thickness")
        width = data.get("width")
        part_no = data.get("part_no")
        make = data.get("make")
        user = session.get("username")
        barcode = f"{width}x{type1}:{part_no}"
        materialType = f"{width}x{type1}:{part_no}"

        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "error": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Check if material exists
        cursor.execute(
            "SELECT COUNT(*) FROM Alu_Raw_Materials WHERE material_Id = ?",
            (material_Id,),
        )
        if cursor.fetchone()[0] == 0:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "Material with the given ID does not exist.",
                    }
                ),
                404,
            )

        # Perform update
        cursor.execute(
            """
            UPDATE Alu_Raw_Materials
            SET thickness = ?, width = ?, part_no = ?, make = ?, [user] = ?, materialType = ?, barcode = ?
            WHERE material_Id = ?
        """,
            (type1, width, part_no, make, user, materialType, barcode, material_Id),
        )
        conn.commit()

        if cursor.rowcount == 0:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "No changes were made to the raw material.",
                    }
                ),
                400,
            )

        return jsonify(
            {"success": True, "message": "Raw material updated successfully."}
        )

    except pyodbc.IntegrityError as e:
        if "UNIQUE constraint failed: Alu_Raw_Materials.materialType" in str(e):
            return (
                jsonify({"success": False, "error": "Material Type must be unique."}),
                409,
            )
        return (
            jsonify({"success": False, "error": f"Database integrity error: {str(e)}"}),
            400,
        )

    except pyodbc.Error as e:
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500

    except Exception as e:
        return (
            jsonify(
                {"success": False, "error": f"An unexpected error occurred: {str(e)}"}
            ),
            500,
        )

    finally:
        if conn:
            conn.close()


@app.route("/h-raw-material", methods=["GET", "POST"])
def raw_material2():
    """Handle raw material addition (POST) and retrieval with pagination (GET)."""

    if "username" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        try:
            # Extract required fields from form
            depth = request.form["depth"]
            width = request.form["width"]
            height = request.form["height"]
            part_no = request.form["part_no"]
            if not depth or not width or not part_no or not height:
                return jsonify({"error": "All fields are required"}), 400

            # Auto-generate fields
            make = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            user = session.get("username")
            # barcode = f"-{type_code}-{lot_no}"  # Example barcode
            material_type = f"{width}x{height}:{part_no}"
            barcode = material_type

            # Connect to MSSQL database
            conn = get_db_connection()
            if not conn:
                return (
                    jsonify(
                        {"success": False, "message": "Database connection failed"}
                    ),
                    500,
                )

            cursor = conn.cursor()

            # Insert data into the Raw_Materials table
            cursor.execute(
                """
                INSERT INTO h_Raw_Materials (length, width,height,part_no, materialType, make, [user], barcode) 
                VALUES (?, ?, ?, ?,?, ?, ?,?)""",
                (depth, width, height, part_no, material_type, make, user, barcode),
            )
            conn.commit()

            return jsonify({"success": True, "message": "Material added successfully!"})

        except pyodbc.Error as e:
            print(f"Database Error: {e}")
            return (
                jsonify({"success": False, "message": f"Database error: {str(e)}"}),
                500,
            )

        finally:
            conn.close()

    # Handle GET request for pagination
    try:
        current_page = int(request.args.get("page", 1))
        limit = int(request.args.get("limit", 10))
        offset = (current_page - 1) * limit

        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "message": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Fetch paginated raw materials
        cursor.execute(
            """
            SELECT * FROM h_Raw_Materials 
            ORDER BY material_Id
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
        """,
            (offset, limit),
        )
        raw_materials = cursor.fetchall()

        # Calculate total pages
        cursor.execute("SELECT COUNT(*) FROM h_Raw_Materials")
        total_rows = cursor.fetchone()[0]
        total_pages = (total_rows + limit - 1) // limit

        conn.close()

        # Render the template with paginated data
        return render_template(
            "housing_raw_material.html",
            raw_materials=raw_materials,
            page=current_page,
            total_pages=total_pages,
            limit=limit,
        )

    except pyodbc.Error as e:
        print(f"Database Error: {e}")
        return jsonify({"success": False, "message": f"Database error: {str(e)}"}), 500


@app.route("/delete-h_raw-material/<int:raw_material_id>", methods=["DELETE"])
def delete_raw_material2(raw_material_id):
    """Delete a raw material entry from MSSQL database."""

    try:
        print(f"Attempting to delete alu_raw material with ID: {raw_material_id}")

        # Check if user is logged in
        if "username" not in session:
            print("Unauthorized access attempt.")
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        # Connect to the MSSQL database
        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "error": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Check if the material exists
        cursor.execute(
            "SELECT * FROM h_Raw_Materials WHERE material_Id = ?", (raw_material_id,)
        )
        row = cursor.fetchone()

        if row is None:
            print(f"Material with ID {raw_material_id} not found.")
            return jsonify({"success": False, "error": "Raw material not found"}), 404

        # Delete the raw material
        cursor.execute(
            "DELETE FROM h_Raw_Materials WHERE material_Id = ?", (raw_material_id,)
        )
        conn.commit()

        if cursor.rowcount == 0:
            print(f"Failed to delete material with ID {raw_material_id}.")
            return (
                jsonify({"success": False, "error": "Failed to delete material"}),
                500,
            )

        print(f"Successfully deleted material with ID {raw_material_id}.")
        return jsonify({"success": True})

    except pyodbc.Error as e:
        print(f"Database Error: {e}")
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500

    finally:
        if conn:
            conn.close()


@app.route("/update-h_raw-material/<int:material_Id>", methods=["POST"])
def update_raw_material2(material_Id):
    """Update raw material details in MSSQL database."""
    conn = None  # Ensure conn is defined before try

    try:
        data = request.json
        print(data)

        # Validate required fields
        required_fields = [
            "depth",
            "width",
            "height",
            "part_no",
            "make",
            "user",
            "barcode",
        ]
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"Missing fields: {', '.join(missing_fields)}",
                    }
                ),
                400,
            )

        depth = data.get("depth")
        width = data.get("width")
        height = data.get("height")
        part_no = data.get("part_no")
        make = data.get("make")
        user = session.get("username")
        barcode = f"{width}x{height}:{part_no}"
        materialType = f"{width}x{height}:{part_no}"

        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "error": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Check if material exists
        cursor.execute(
            "SELECT COUNT(*) FROM h_Raw_Materials WHERE material_Id = ?", (material_Id,)
        )
        if cursor.fetchone()[0] == 0:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "Material with the given ID does not exist.",
                    }
                ),
                404,
            )

        # Perform update
        cursor.execute(
            """
            UPDATE h_Raw_Materials
            SET length = ?, width = ?, height = ?, part_no = ?, make = ?, [user] = ?, materialType = ?, barcode = ?
            WHERE material_Id = ?
        """,
            (
                depth,
                width,
                height,
                part_no,
                make,
                user,
                materialType,
                barcode,
                material_Id,
            ),
        )
        conn.commit()

        if cursor.rowcount == 0:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "No changes were made to the raw material.",
                    }
                ),
                400,
            )

        return jsonify(
            {"success": True, "message": "Raw material updated successfully."}
        )

    except pyodbc.IntegrityError as e:
        if "UNIQUE constraint failed: h_Raw_Materials.materialType" in str(e):
            return (
                jsonify({"success": False, "error": "Material Type must be unique."}),
                409,
            )
        return (
            jsonify({"success": False, "error": f"Database integrity error: {str(e)}"}),
            400,
        )

    except pyodbc.Error as e:
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500

    except Exception as e:
        return (
            jsonify(
                {"success": False, "error": f"An unexpected error occurred: {str(e)}"}
            ),
            500,
        )

    finally:
        if conn:
            conn.close()


@app.route("/delete-recipe/<int:recipe_id>", methods=["DELETE"])
def delete_recipe(recipe_id):
    """Delete a recipe from MSSQL database."""

    try:
        # Debugging: Check if session contains username
        if "username" not in session:
            print("Session does not contain 'username'. Current session:", session)
            return jsonify({"success": False, "error": "Unauthorized access"}), 401

        # Connect to MSSQL database
        conn = get_db_connection()
        if not conn:
            return (
                jsonify({"success": False, "error": "Database connection failed"}),
                500,
            )

        cursor = conn.cursor()

        # Check if the recipe exists before deletion
        cursor.execute("SELECT COUNT(*) FROM Recipe WHERE Recipe_ID = ?", (recipe_id,))
        if cursor.fetchone()[0] == 0:
            return jsonify({"success": False, "error": "Recipe not found"}), 404

        # Delete the recipe
        cursor.execute("DELETE FROM Recipe_Details1 WHERE Recipe_ID = ?", (recipe_id,))
        cursor.execute(
            "DELETE FROM Inspection_Settings WHERE Recipe_ID = ?", (recipe_id,)
        )
        cursor.execute("DELETE FROM Sub_Menu WHERE Recipe_ID = ?", (recipe_id,))
        cursor.execute("DELETE FROM Recipe WHERE Recipe_ID = ?", (recipe_id,))
        conn.commit()

        if cursor.rowcount == 0:
            return jsonify({"success": False, "error": "Failed to delete recipe"}), 500

        return jsonify({"success": True, "message": "Recipe deleted successfully"})

    except pyodbc.Error as e:
        return jsonify({"success": False, "error": f"Database error: {str(e)}"}), 500

    except Exception as e:
        print(f"Error while deleting recipe: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        if conn:
            conn.close()


def fetch_data_as_dict(cursor):
    """Helper function to convert query result to dictionary."""
    columns = [column[0] for column in cursor.description]
    result = []
    for row in cursor.fetchall():
        result.append(dict(zip(columns, row)))
    return result

@app.route("/export_recipe/<int:recipe_id>")
def export_recipe(recipe_id):
    conn = None
    try:
        # ================== DB Connection ==================
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "danger")
            return redirect(url_for("recipe_details", recipe_id=recipe_id))
        cursor = conn.cursor()

        # ================== Fetch Recipe Details ==================
        cursor.execute(
            "SELECT * FROM Recipe_Details1 WHERE Recipe_ID = ?", (recipe_id,)
        )
        recipe_details = cursor.fetchone()  # One row expected
        recipe_columns = (
            [column[0] for column in cursor.description] if cursor.description else []
        )

        # ================== Fetch Sub Menu ==================
        cursor.execute("SELECT * FROM Sub_Menu WHERE Recipe_ID = ?", (recipe_id,))
        sub_menu = cursor.fetchall()
        sub_menu_columns = (
            [column[0] for column in cursor.description] if cursor.description else []
        )

        # ================== Fetch Inspection Settings ==================
        cursor.execute(
            "SELECT * FROM Inspection_Settings WHERE Recipe_ID = ?", (recipe_id,)
        )
        inspection_menu = cursor.fetchall()
        inspection_columns = (
            [column[0] for column in cursor.description] if cursor.description else []
        )

        # ================== Fetch Recipe Name ==================
        cursor.execute(
            "SELECT Recipe_Name FROM Recipe WHERE Recipe_ID = ?", (recipe_id,)
        )
        row = cursor.fetchone()
        Recipe_Name = row[0] if row else "Unknown"

        # ================== Helper: normalize value ==================
        def normalize_val(val):
            return val if val not in ("", None) else None

        def write_vertical(ws, columns, data_rows):
            """Write columns & data in vertical manner"""
            if not columns or not data_rows:
                ws.append(["No data found"])
                return

            for idx, row in enumerate(data_rows, start=1):
                # ws.append(["---- Record {} ----".format(idx), ""])
                for col, val in zip(columns, row):
                    ws.append([col, normalize_val(val)])
                ws.append([])  # empty row as separator

            # Center align
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-fit column width
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_len + 2

        # ================== Create Excel Workbook ==================
        wb = openpyxl.Workbook()

        # ---------- Recipe Details ----------
        ws1 = wb.active
        ws1.title = "Recipe Details"
        if recipe_details:
            write_vertical(ws1, recipe_columns, [recipe_details])
        else:
            ws1.append(["No recipe details found"])

        # ---------- Sub Menu ----------
        ws2 = wb.create_sheet(title="Sub Menu")
        if sub_menu:
            write_vertical(ws2, sub_menu_columns, sub_menu)
        else:
            ws2.append(["No sub menu found"])

        # ---------- Inspection Settings ----------
        ws3 = wb.create_sheet(title="Inspection Settings")
        if inspection_menu:
            write_vertical(ws3, inspection_columns, inspection_menu)
        else:
            ws3.append(["No inspection settings found"])

        # ================== Save to BytesIO ==================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # ================== Sanitize filename ==================
        safe_name = "".join(
            c if c.isalnum() or c in " _-" else "_" for c in Recipe_Name
        )
        filename = f"Recipe-Name_{safe_name}_Recipe-Id_{recipe_id}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        print(f"❌ Error exporting recipe {recipe_id}: {str(e)}")
        flash(f"Error exporting to Excel: {str(e)}", "danger")
        return redirect(url_for("recipe_details", recipe_id=recipe_id))

    finally:
        if conn:
            conn.close()


@app.route("/recipe/<int:recipe_id>")
def recipe_details(recipe_id):
    """Fetch and display recipe details."""
    try:
        # Check user session
        if "username" not in session:
            flash("Please log in to access this page.", "warning")
            return redirect(url_for("login"))

        # Establish database connection
        conn = get_db_connection()
        if not conn:
            flash("Database connection failed.", "danger")
            return redirect(url_for("login"))
        cursor1 = conn.cursor()
        cursor1.execute(
            "SELECT * FROM Recipe_Details1 WHERE Recipe_ID = ?", (recipe_id,)
        )
        recipe_details = cursor1.fetchall()

        cursor2 = conn.cursor()
        cursor2.execute("SELECT * FROM Sub_Menu WHERE Recipe_ID = ?", (recipe_id,))
        sub_menu = cursor2.fetchall()
        print("sub_menu:", sub_menu)

        cursor3 = conn.cursor()
        cursor3.execute(
            "SELECT * FROM Inspection_Settings WHERE Recipe_ID = ?", (recipe_id,)
        )
        inspection_menu = cursor3.fetchall()
        print("inspection_menu:", inspection_menu)

        cursor4 = conn.cursor()
        cursor4.execute(
            "SELECT Recipe_Name FROM Recipe WHERE Recipe_ID = ?", (recipe_id,)
        )
        row = cursor4.fetchone()  # fetchone returns a single tuple, e.g. ('M5',)
        Recipe_Name = row[0] if row else "Unknown"  # extract the string

        # cursor3 = conn.cursor()
        # cursor3.execute("SELECT * FROM Insection_Settings WHERE Recipe_ID = ?", (recipe_id,))
        # inspection_menu = cursor3.fetchall()

        if not recipe_details:
            flash("Recipe not found!", "danger")
            return redirect(url_for("index"))

        # Process data safely
        pos_values = extract_pos_values(recipe_details) if recipe_details else []
        submenu_values = extract_submenu_vaues(sub_menu) if sub_menu else []
        barcode_value = "S3"  # Replace with actual barcode retrieval logic

        flash(
            f"POS values written to PLC successfully for Recipe ID {recipe_id}.",
            "success",
        )

        return render_template(
            "recipe_details.html",
            recipe_id=recipe_id,
            Recipe_Name=Recipe_Name,
            recipe_details=recipe_details,
            sub_menu=sub_menu[0],
            inspection_menu=inspection_menu[0],
            pos_values=pos_values,  # Assuming last value should be removed
            barcode_value=barcode_value,
        )

    except pyodbc.Error as e:
        flash(f"Database error: {str(e)}", "danger")
        print(f"Database error at pyodbc error: {str(e)}", "danger")

        return redirect(url_for("index"))

    except Exception as e:
        flash(f"An unexpected error occurred: {str(e)}", "danger")
        print(f"An unexpected error occurred at exception: {str(e)}", "danger")
        return redirect(url_for("index"))

    finally:
        if conn:
            conn.close()


def extract_pos_values(recipe_details):
    if not recipe_details:
        return []

    # Assuming the structure of each tuple in `recipe_details`
    first_row = recipe_details[0]  # Get the first tuple

    pos_values = [
        first_row[1],  # Assuming `pos1` is at index 1
        first_row[2],  # Assuming `pos2` is at index 2
        first_row[3],  # Assuming `pos3` is at index 3
        first_row[4],  # Assuming `pos4` is at index 4
        first_row[5],  # Assuming `pos5` is at index 5
        first_row[6],  # Assuming `pos6` is at index 6
        first_row[7],  # Assuming `pos7` is at index 7
        first_row[8],  # Assuming `pos8` is at index 8
        first_row[9],  # Assuming `pos9` is at index 9
        # first_row[10],  # Assuming `recipe_id` is at index 10
    ]

    return pos_values


def extract_submenu_vaues(sub_menu):
    if not sub_menu:
        return []

    first_row = sub_menu[0]  # Get the first tuple

    submenu_values = [
        # first_row[1],  # Assuming `motor_speed` is at index 1
        # first_row[2],  # Assuming `motor_stroke` is at index 2
        # first_row[3],  # Assuming `other_Speed_force` is at index 3
        # first_row[4],  # Assuming `alu_coil_width` is at index 4
        first_row[5],
        first_row[6],
        first_row[7],
        first_row[8],
        first_row[9],
        first_row[19],
        first_row[11],
        first_row[12],
        first_row[13],
        first_row[14],
        first_row[15],
        first_row[16],
        first_row[17],
        first_row[18],
        first_row[19],
        first_row[20],
        first_row[21],
        first_row[22],
        first_row[23],
        first_row[24],
        first_row[25],
        first_row[26],
    ]
    return submenu_values


BARCODE_NODE_ID = f'ns=3;s="dbRecipe1"."qrCodeDecoilder"'


def fetch_barcode_from_opcua():
    try:
        client = Client(ENDPOINT_URL)
        client.connect()
        barcode_value = client.get_node(BARCODE_NODE_ID).get_value()
        client.disconnect()
        return barcode_value
    except Exception as e:
        return None  # Handle errors gracefully


@app.route("/compare-pos", methods=["POST"])
def compare_pos():
    data = request.json  # Receive data from frontend
    pos_value = data.get("posValue")

    if pos_value is None:
        return jsonify({"error": "Invalid data"}), 400

    # Fetch barcode value from OPC UA server fetch_barcode_from_opcua()
    barcode_value = fetch_barcode_from_opcua()
    print("Barcode_value", barcode_value)
    print("Pos_Value", pos_value)
    if barcode_value is None:
        return jsonify({"error": "Failed to fetch barcode from OPC UA"}), 500

    # Comparison logic
    match = pos_value == barcode_value

    return jsonify({"match": match})


@app.route("/update_recipe", methods=["POST"])
def update_recipe():
    """Update an existing recipe with new details."""

    # Ensure user is logged in
    if "username" not in session:
        flash("Please log in to update recipes.", "warning")
        return redirect(url_for("login"))

    # -------------------------
    # Extract recipe info
    # -------------------------
    recipe_id = request.form.get("recipe_id")
    recipe_name = request.form.get("recipe_name")
    filter_size = request.form.get("filter_size")
    filter_code = request.form.get("filter_code")
    art_no = request.form.get("art_no")

    # Helper for numeric values
    def clean_numeric(value):
        if value is None or value.strip() == "":
            return None
        return value

    # -------------------------
    # Extract fields (same as add_recipe)
    # -------------------------
    pleat_height = clean_numeric(request.form.get("Pleat_Height"))
    soft_touch = clean_numeric(request.form.get("Soft_Touch"))
    feeder1_media_thickness = clean_numeric(request.form.get("Feeder1_Media_Thickness"))
    feeder1_park_position = clean_numeric(request.form.get("Feeder1_Park_Position"))
    foil1_length = clean_numeric(request.form.get("Foil1_Length"))
    foil1_length_offset = clean_numeric(request.form.get("Foil1_Length_Offset"))
    foil1_width = clean_numeric(request.form.get("Foil1_Width"))
    puller_start_position = clean_numeric(request.form.get("Puller_Start_Position"))
    autofeeding_offset = clean_numeric(request.form.get("Autofeeding_Offset"))
    filter_box_height = clean_numeric(request.form.get("Filter_Box_Height"))
    filter_box_length = clean_numeric(request.form.get("Filter_Box_Length"))
    pleat_pitch = clean_numeric(request.form.get("Pleat_Pitch"))
    pleat_counts = clean_numeric(request.form.get("Pleat_Counts"))
    foil_low_diameter = clean_numeric(request.form.get("Foil_Low_Diameter"))
    feeding_conveyor_speed = clean_numeric(request.form.get("Feeding_Conveyor_Speed"))
    pack_transfer_rev_position = clean_numeric(
        request.form.get("Pack_Transfer_Rev_Position")
    )
    foil1_tension_set_point = clean_numeric(request.form.get("Foil1_Tension_Set_Point"))
    foil2_tension_set_point = clean_numeric(request.form.get("Foil2_Tension_Set_Point"))
    blade_opening = clean_numeric(request.form.get("Blade_Opening"))
    press_touch = clean_numeric(request.form.get("Press_Touch"))
    feeder2_media_thickness = clean_numeric(request.form.get("Feeder2_Media_Thickness"))
    feeder2_park_position = clean_numeric(request.form.get("Feeder2_Park_Position"))
    foil2_length = clean_numeric(request.form.get("Foil2_Length"))
    foil2_length_offset = clean_numeric(request.form.get("Foil2_Length_Offset"))
    foil2_width = clean_numeric(request.form.get("Foil2_Width"))
    puller_end_position = clean_numeric(request.form.get("Puller_End_Position"))
    puller2_feed_correction = clean_numeric(request.form.get("Puller2_Feed_Correction"))
    puller_extra_stroke_enable = clean_numeric(
        request.form.get("Puller_Extra_Stroke_Enable")
    )
    filter_box_width = clean_numeric(request.form.get("Filter_Box_Width"))
    lid_placement_enable = clean_numeric(request.form.get("Lid_Placement_Enable"))
    lid_placement_position = clean_numeric(request.form.get("Lid_Placement_Position"))
    sync_table_start_position = clean_numeric(
        request.form.get("Sync_Table_Start_Position")
    )
    batch_count = clean_numeric(request.form.get("Batch_Count"))
    discharge_conveyor_speed = clean_numeric(
        request.form.get("Discharge_Conveyor_Speed")
    )
    pack_transfer_park_position = clean_numeric(
        request.form.get("Pack_Transfer_Park_Position")
    )
    media_tension_set_point = clean_numeric(request.form.get("Media_Tension_Set_Point"))

    # Extra Inspection
    databaseAvailable = request.form.get("databaseAvailable")
    Width = request.form.get("Width")
    Height = request.form.get("Height")
    Depth = request.form.get("Depth")
    Art_Num = request.form.get("Art_Num")
    Air_Flow_Set = request.form.get("Air_Flow_Set")
    Pressure_Drop_Setpoint = request.form.get("Pressure_Drop_Setpoint")
    Lower_Tolerance1 = request.form.get("Lower_Tolerance1")
    Lower_Tolerance2 = request.form.get("Lower_Tolerance2")
    Upper_Tolerance1 = request.form.get("Upper_Tolerance1")
    Upper_Tolerance2 = request.form.get("Upper_Tolerance2")
    # lower_fan_speed = request.form.get('Lower_Fan_Speed')
    # upper_fan_speed = request.form.get('Upper_Fan_Speed')
    alu_value = request.form.get("alu_mat", None)
    house_value = request.form.get("house_mat", None)

    # POS values
    pos_values = [request.form.get(f"pos{i}", None) for i in range(1, 10)]

    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        # Update Recipe
        cursor.execute(
            """UPDATE Recipe 
               SET Recipe_Name = ?, Filter_Size = ?, Filter_Code = ?, Art_No = ?
               WHERE Recipe_ID = ?""",
            (recipe_name, filter_size, filter_code, art_no, recipe_id),
        )

        # Update Recipe_Details1
        cursor.execute(
            """UPDATE Recipe_Details1 
               SET Pos1=?, Pos2=?, Pos3=?, Pos4=?, Pos5=?, Pos6=?, Pos7=?, Pos8=?, Pos9=?, Alu_Material=?, House_Material=? 
               WHERE Recipe_ID = ?""",
            (*pos_values, alu_value, house_value, recipe_id),
        )

        # Update Sub_Menu
        cursor.execute(
            """UPDATE Sub_Menu 
               SET Pleat_Height=?, Soft_Touch=?, Feeder1_Media_Thickness=?, Feeder1_Park_Position=?,
                   Foil1_Length=?, Foil1_Length_Offset=?, Foil1_Width=?, Puller_Start_Position=?, Autofeeding_Offset=?,
                   Filter_Box_Height=?, Filter_Box_Length=?, Pleat_Pitch=?, Pleat_Counts=?, Foil_Low_Diameter=?,
                   Feeding_Conveyor_Speed=?, Pack_Transfer_Rev_Position=?, Foil1_Tension_Set_Point=?, Foil2_Tension_Set_Point=?,
                   Blade_Opening=?, Press_Touch=?, Feeder2_Media_Thickness=?, Feeder2_Park_Position=?,
                   Foil2_Length=?, Foil2_Length_Offset=?, Foil2_Width=?, Puller_End_Position=?,
                   Puller2_Feed_Correction=?, Puller_Extra_Stroke_Enable=?, Filter_Box_Width=?,
                   Lid_Placement_Enable=?, Lid_Placement_Position=?, Sync_Table_Start_Position=?,
                   Batch_Count=?, Discharge_Conveyor_Speed=?, Pack_Transfer_Park_Position=?, Media_Tension_Set_Point=? 
               WHERE Recipe_ID = ?""",
            (
                pleat_height,
                soft_touch,
                feeder1_media_thickness,
                feeder1_park_position,
                foil1_length,
                foil1_length_offset,
                foil1_width,
                puller_start_position,
                autofeeding_offset,
                filter_box_height,
                filter_box_length,
                pleat_pitch,
                pleat_counts,
                foil_low_diameter,
                feeding_conveyor_speed,
                pack_transfer_rev_position,
                foil1_tension_set_point,
                foil2_tension_set_point,
                blade_opening,
                press_touch,
                feeder2_media_thickness,
                feeder2_park_position,
                foil2_length,
                foil2_length_offset,
                foil2_width,
                puller_end_position,
                puller2_feed_correction,
                puller_extra_stroke_enable,
                filter_box_width,
                lid_placement_enable,
                lid_placement_position,
                sync_table_start_position,
                batch_count,
                discharge_conveyor_speed,
                pack_transfer_park_position,
                media_tension_set_point,
                recipe_id,
            ),
        )

        # Update Inspection_Settings
        cursor.execute(
            """UPDATE Inspection_Settings 
               SET databaseAvailable=?, Width=?, Height=?, Depth=?, Art_No=?, Air_Flow_Set=?, Pressure_Drop_Setpoint=?, 
                   Lower_Tolerance1=?, Lower_Tolerance2=?, Upper_Tolerance1=?, Upper_Tolerance2=?,Lower_Fan_Speed=?,Upper_Fan_Speed=? 
               WHERE Recipe_ID=?""",
            (
                databaseAvailable,
                Width,
                Height,
                Depth,
                Art_Num,
                Air_Flow_Set,
                Pressure_Drop_Setpoint,
                Lower_Tolerance1,
                Lower_Tolerance2,
                Upper_Tolerance1,
                Upper_Tolerance2,
                recipe_id,
            ),
        )

        conn.commit()
        flash("Recipe updated successfully!", "success")

    except pyodbc.Error as e:
        conn.rollback()
        flash(f"Database error: {str(e)}", "danger")
    except Exception as e:
        conn.rollback()
        flash(f"Unexpected error: {str(e)}", "danger")
    finally:
        conn.close()

    return redirect(url_for("recipe_list"))


@app.route("/add_recipe", methods=["POST"])
def add_recipe():
    """Add a new recipe to the database."""
    data = request.form.to_dict()
    print("Form data received:", data)  # Debug

    # Ensure user is logged in
    if "username" not in session:
        flash("Please log in to add recipes.", "warning")
        return redirect(url_for("login"))

    # -------------------------
    # Helper to clean numeric inputs
    # -------------------------
    def clean_numeric(value):
        if value is None or value.strip() == "":
            return None
        return value

    # -------------------------
    # Extract main recipe info
    # -------------------------
    recipe_id = request.form.get("recipe_id")
    recipe_name = request.form.get("recipe_name")
    filter_size = request.form.get("filter_size")
    filter_code = request.form.get("filter_code")
    art_no = request.form.get("art_no")

    # -------------------------
    # Machine Settings Fields (cleaned)
    # -------------------------
    pleat_height = clean_numeric(request.form.get("Pleat_Height"))
    soft_touch = clean_numeric(request.form.get("Soft_Touch"))
    feeder1_media_thickness = clean_numeric(request.form.get("Feeder1_Media_Thickness"))
    feeder1_park_position = clean_numeric(request.form.get("Feeder1_Park_Position"))
    foil1_length = clean_numeric(request.form.get("Foil1_Length"))
    foil1_length_offset = clean_numeric(request.form.get("Foil1_Length_Offset"))
    foil1_width = clean_numeric(request.form.get("Foil1_Width"))
    puller_start_position = clean_numeric(request.form.get("Puller_Start_Position"))
    autofeeding_offset = clean_numeric(request.form.get("Autofeeding_Offset"))
    filter_box_height = clean_numeric(request.form.get("Filter_Box_Height"))
    filter_box_length = clean_numeric(request.form.get("Filter_Box_Length"))
    pleat_pitch = clean_numeric(request.form.get("Pleat_Pitch"))
    pleat_counts = clean_numeric(request.form.get("Pleat_Counts"))
    foil_low_diameter = clean_numeric(request.form.get("Foil_Low_Diameter"))
    feeding_conveyor_speed = clean_numeric(request.form.get("Feeding_Conveyor_Speed"))
    pack_transfer_rev_position = clean_numeric(
        request.form.get("Pack_Transfer_Rev_Position")
    )
    foil1_tension_set_point = clean_numeric(request.form.get("Foil1_Tension_Set_Point"))
    foil2_tension_set_point = clean_numeric(request.form.get("Foil2_Tension_Set_Point"))
    blade_opening = clean_numeric(request.form.get("Blade_Opening"))
    press_touch = clean_numeric(request.form.get("Press_Touch"))
    feeder2_media_thickness = clean_numeric(request.form.get("Feeder2_Media_Thickness"))
    feeder2_park_position = clean_numeric(request.form.get("Feeder2_Park_Position"))
    foil2_length = clean_numeric(request.form.get("Foil2_Length"))
    foil2_length_offset = clean_numeric(request.form.get("Foil2_Length_Offset"))
    foil2_width = clean_numeric(request.form.get("Foil2_Width"))
    puller_end_position = clean_numeric(request.form.get("Puller_End_Position"))
    puller2_feed_correction = clean_numeric(request.form.get("Puller2_Feed_Correction"))
    puller_extra_stroke_enable = clean_numeric(
        request.form.get("Puller_Extra_Stroke_Enable")
    )
    filter_box_width = clean_numeric(request.form.get("Filter_Box_Width"))
    lid_placement_enable = clean_numeric(request.form.get("Lid_Placement_Enable"))
    lid_placement_position = clean_numeric(request.form.get("Lid_Placement_Position"))
    sync_table_start_position = clean_numeric(
        request.form.get("Sync_Table_Start_Position")
    )
    batch_count = clean_numeric(request.form.get("Batch_Count"))
    discharge_conveyor_speed = clean_numeric(
        request.form.get("Discharge_Conveyor_Speed")
    )
    pack_transfer_park_position = clean_numeric(
        request.form.get("Pack_Transfer_Park_Position")
    )
    media_tension_set_point = clean_numeric(request.form.get("Media_Tension_Set_Point"))

    # -------------------------
    # Extra Inspection Fields
    # -------------------------
    databaseAvailable = request.form.get("databaseAvailable")
    Width = request.form.get("Width")
    Height = request.form.get("Height")
    Depth = request.form.get("Depth")
    Art_Num = request.form.get("Art_Num")
    Air_Flow_Set = request.form.get("Air_Flow_Set")
    Pressure_Drop_Setpoint = request.form.get("Pressure_Drop_Setpoint")
    Lower_Tolerance1 = request.form.get("Lower_Tolerance1")
    Lower_Tolerance2 = request.form.get("Lower_Tolerance2")
    Upper_Tolerance1 = request.form.get("Upper_Tolerance1")
    Upper_Tolerance2 = request.form.get("Upper_Tolerance2")
    # lower_fan_speed = request.form.get('Lower_Fan_Speed')
    # upper_fan_speed = request.form.get('Upper_Fan_Speed')
    alu_value = request.form.get("alu_mat", None)
    house_value = request.form.get("house_mat", None)

    # POS values
    pos_values = [request.form.get(f"pos{i}", None) for i in range(1, 10)]

    conn = get_db_connection()
    try:
        cursor = conn.cursor()

        print("👉 Inserting into Recipe table...")
        cursor.execute(
            """INSERT INTO Recipe (Recipe_ID, Recipe_Name, Filter_Size, Filter_Code, Art_No) 
               VALUES (?, ?, ?, ?, ?)""",
            (recipe_id, recipe_name, filter_size, filter_code, art_no),
        )
        print("✅ Inserted into Recipe")

        print("👉 Inserting into Recipe_Details1 table...")
        cursor.execute(
            """INSERT INTO Recipe_Details1 
               (Recipe_ID, Pos1, Pos2, Pos3, Pos4, Pos5, Pos6, Pos7, Pos8, Pos9, Alu_Material, House_Material) 
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (recipe_id, *pos_values, alu_value, house_value),
        )
        print("✅ Inserted into Recipe_Details1")

        print("👉 Inserting into Sub_Menu table...")
        cursor.execute(
            """INSERT INTO Sub_Menu 
               (Recipe_ID, Pleat_Height, Soft_Touch, Feeder1_Media_Thickness, Feeder1_Park_Position,
                Foil1_Length, Foil1_Length_Offset, Foil1_Width, Puller_Start_Position, Autofeeding_Offset,
                Filter_Box_Height, Filter_Box_Length, Pleat_Pitch, Pleat_Counts, Foil_Low_Diameter,
                Feeding_Conveyor_Speed, Pack_Transfer_Rev_Position, Foil1_Tension_Set_Point, Foil2_Tension_Set_Point,
                Blade_Opening, Press_Touch, Feeder2_Media_Thickness, Feeder2_Park_Position,
                Foil2_Length, Foil2_Length_Offset, Foil2_Width, Puller_End_Position,
                Puller2_Feed_Correction, Puller_Extra_Stroke_Enable, Filter_Box_Width,
                Lid_Placement_Enable, Lid_Placement_Position, Sync_Table_Start_Position,
                Batch_Count, Discharge_Conveyor_Speed, Pack_Transfer_Park_Position, Media_Tension_Set_Point) 
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                recipe_id,
                pleat_height,
                soft_touch,
                feeder1_media_thickness,
                feeder1_park_position,
                foil1_length,
                foil1_length_offset,
                foil1_width,
                puller_start_position,
                autofeeding_offset,
                filter_box_height,
                filter_box_length,
                pleat_pitch,
                pleat_counts,
                foil_low_diameter,
                feeding_conveyor_speed,
                pack_transfer_rev_position,
                foil1_tension_set_point,
                foil2_tension_set_point,
                blade_opening,
                press_touch,
                feeder2_media_thickness,
                feeder2_park_position,
                foil2_length,
                foil2_length_offset,
                foil2_width,
                puller_end_position,
                puller2_feed_correction,
                puller_extra_stroke_enable,
                filter_box_width,
                lid_placement_enable,
                lid_placement_position,
                sync_table_start_position,
                batch_count,
                discharge_conveyor_speed,
                pack_transfer_park_position,
                media_tension_set_point,
            ),
        )
        print("✅ Inserted into Sub_Menu")

        print("👉 Inserting into Inspection_Settings table...")
        cursor.execute(
            """INSERT INTO Inspection_Settings 
               (Recipe_ID, databaseAvailable, Width, Height, Depth, Art_No, Air_Flow_Set, Pressure_Drop_Setpoint, 
                Lower_Tolerance1, Lower_Tolerance2, Upper_Tolerance1, Upper_Tolerance2) 
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                recipe_id,
                databaseAvailable,
                Width,
                Height,
                Depth,
                Art_Num,
                Air_Flow_Set,
                Pressure_Drop_Setpoint,
                Lower_Tolerance1,
                Lower_Tolerance2,
                Upper_Tolerance1,
                Upper_Tolerance2,
            ),
        )
        print("✅ Inserted into Inspection_Settings")

        conn.commit()
        print("🎉 All inserts committed successfully!")
        flash("Recipe added successfully!", "success")

    except pyodbc.Error as e:
        print(f"❌ Database error occurred: {str(e)}")
        conn.rollback()
        flash(f"Database error: {str(e)}", "danger")

    except Exception as e:
        print(f"❌ Unexpected error occurred: {str(e)}")
        conn.rollback()
        flash(f"Unexpected error: {str(e)}", "danger")

    finally:
        conn.close()
        print("🔒 Connection closed.")

    return redirect(url_for("recipe_list"))


def update_plc_with_values(pos_values, submenu_values, recipe_info):
    """
    Update PLC with POS values, Recipe Name, Recipe ID, and submenu values
    using the latest PLC tag mapping.
    """
    from opcua import Client, ua

    client = Client(ENDPOINT_URL)
    client.session_timeout = 30000  # Adjust timeout as needed

    try:
        client.connect()

        # Extract Recipe Name and Recipe ID from recipe_info
        recipe_name = recipe_info[3]  # 'Recipe6'
        recipe_id = recipe_info[4]  # 6

        # ✅ Write POS values (expects dict)
        for i in range(1, 10):
            key = f"Pos{i}"
            if key not in pos_values:
                print(f"⚠️ Missing {key} in POS values, skipping...")
                continue

            pos_value = pos_values[key]
            node_id1 = f'ns=3;s="dbRecipe1"."Recipe"."decoilerSelection"[{i}]'
            node_id2 = f'ns=3;s="dbRecipe1"."Recipe"."decolerRoll"[{i}]'

            try:
                node1 = client.get_node(node_id1)
                node1.set_value(
                    ua.DataValue(ua.Variant(bool(pos_value), ua.VariantType.Boolean))
                )

                node2 = client.get_node(node_id2)
                node2.set_value(
                    ua.DataValue(ua.Variant(str(pos_value), ua.VariantType.String))
                )

                print(f"✔️ POS[{i}] = {pos_value} written successfully.")

            except Exception as node_error:
                print(f"❌ Error processing POS[{i}] Node: {node_error}")

        # ✅ Write Recipe Name and Recipe ID
        try:
            recipe_name_node = client.get_node(
                'ns=3;s="dbRecipe1"."Recipe"."RecipeName"'
            )
            recipe_id_node = client.get_node('ns=3;s="dbRecipe1"."Recipe"."RecipeID"')

            recipe_name_node.set_value(
                ua.DataValue(ua.Variant(recipe_name, ua.VariantType.String))
            )
            recipe_id_node.set_value(
                ua.DataValue(ua.Variant(recipe_id, ua.VariantType.Int32))
            )

            print(
                f"✔️ Recipe Name ({recipe_name}) and Recipe ID ({recipe_id}) written to PLC."
            )
        except Exception as recipe_error:
            print(f"❌ Error writing Recipe Name and Recipe ID: {recipe_error}")

        # ✅ Updated submenu tag mapping
        db_to_plc_map = {
            "Pleat_Height": "Pleat_Height",
            "Soft_Touch": "Soft_Touch",
            "Feeder1_Media_Thickness": "Left_Blade_MediaTHickness",
            "Feeder1_Park_Position": "Left_Blade_Start_pos",
            "Foil1_Length": "Corr1Feed_length",
            "Foil1_Length_Offset": "Corr1feed_length_offset",
            "Foil1_Width": "Corr1feed_length_width",
            "Puller_Start_Position": "Puller_Start_pos",
            "Autofeeding_Offset": "Auto_Feeding_Offset",
            "Filter_Box_Height": "Filter_Box_Height",
            "Filter_Box_Length": "Filter_Box_Length",
            "Pleat_Pitch": "Set_Pleat_Pitch",
            "Pleat_Counts": "Set_Pleat_Count",
            "Foil_Low_Diameter": "Low_Dia_Set",
            "Feeding_Conveyor_Speed": "Feeding_Conveyor_Speed",
            "Pack_Transfer_Rev_Position": "Filter_Transfer_pos_rev",
            "Foil1_Tension_Set_Point": "Corr_1_Set_point",
            "Foil2_Tension_Set_Point": "Corr_2_Set_point",
            "Blade_Opening": "Blade_opening",
            "Press_Touch": "Press_Touch",
            "Feeder2_Media_Thickness": "Right_Blade_MediaThickness",
            "Feeder2_Park_Position": "Right_Blade_Start_pos",
            "Foil2_Length": "Corr2feed_length",
            "Foil2_Length_Offset": "Corr2feed_length_offset",
            "Foil2_Width": "Corr2feed_length_width",
            "Puller_End_Position": "Puller_End_pos",
            "Puller2_Feed_Correction": "Puller2_Feed_Correction",
            "Puller_Extra_Stroke_Enable": "PULLER_Extra stoke_Enable",
            "Filter_Box_Width": "Filter_Box_Width",
            "Lid_Placement_Enable": "Lid_Placement_Enable",
            "Lid_Placement_Position": "Lid_Placement_Pos",
            "Sync_Table_Start_Position": "Sync_Table_Start_Counts",
            "Batch_Count": "Set_Batch_COunt",
            "Discharge_Conveyor_Speed": "Discharge_COnvyor_Speed",
            "Pack_Transfer_Park_Position": "Filter_Transfer_pos_fwd",
            "Media_Tension_Set_Point": "Media_puller_2_SP",
        }

        # ✅ Write submenu values
        for db_field, plc_tag in db_to_plc_map.items():
            if db_field not in submenu_values:
                print(f"⚠️ Missing {db_field} in provided values, skipping...")
                continue

            value = submenu_values[db_field]
            submenu_node_id = f'ns=3;s="dbRecipe1"."Recipe"."{plc_tag}"'

            try:
                node = client.get_node(submenu_node_id)
                data_type = node.get_data_type_as_variant_type()

                if data_type == ua.VariantType.Float:
                    plc_value = float(value)
                elif data_type == ua.VariantType.Int32:
                    plc_value = int(value)
                elif data_type == ua.VariantType.Boolean:
                    plc_value = bool(value)
                else:
                    plc_value = str(value)

                node.set_value(ua.DataValue(ua.Variant(plc_value, data_type)))
                print(f"✔️ Wrote {db_field} ({plc_tag}) = {plc_value}")

            except Exception as submenu_error:
                print(f"❌ Error writing {db_field} ({plc_tag}): {submenu_error}")

        print(
            "✅ POS values, Recipe Name, Recipe ID, and all Submenu values written successfully to PLC."
        )

    except Exception as e:
        print(f"❌ Error updating PLC: {e}")

    finally:
        client.disconnect()


@app.route("/start_recipe/<int:recipe_id>", methods=["POST"])
def start_recipe(recipe_id):
    """Start a recipe batch and log the process"""
    if "username" not in session:
        flash("Please log in to start a recipe.", "warning")
        return redirect(url_for("login"))

    conn = get_db_connection()
    if not conn:
        flash("Database connection failed.", "danger")
        return redirect(url_for("recipe_list"))

    try:
        cursor = conn.cursor()

        # Fetch Recipe Info
        cursor.execute(
            "SELECT Filter_Size, Filter_Code, Art_No, Recipe_Name, Recipe_ID "
            "FROM Recipe WHERE Recipe_ID = ?",
            (recipe_id,),
        )
        recipe_info = cursor.fetchone()

        # Fetch Recipe Details
        cursor.execute(
            "SELECT * FROM Recipe_Details1 WHERE Recipe_ID = ?", (recipe_id,)
        )
        recipe_details_row = cursor.fetchone()
        recipe_details = row_to_dict(cursor, recipe_details_row)
        cursor.execute("SELECT * FROM Sub_Menu WHERE Recipe_ID = ?", (recipe_id,))
        sub_menu_row = cursor.fetchone()

        if not recipe_details_row or not sub_menu_row:
            flash("Recipe details not found!", "danger")
            return redirect(url_for("recipe_details", recipe_id=recipe_id))

        # ✅ Convert to dicts

        sub_menu = row_to_dict(cursor, sub_menu_row)

        print("Recipe Details Dict:", recipe_details)
        print("Sub Menu Dict:", sub_menu)
        print("Recipe Info Row:", recipe_info)

        # ✅ Supply dict values to PLC
        update_plc_with_values(recipe_details, sub_menu, recipe_info)

        # Handle JSON payload
        data = request.get_json()
        if not data:
            return jsonify({"error": "Invalid JSON payload"}), 400

        quantity = data.get("quantity")
        if not quantity:
            return jsonify({"error": "Quantity is required"}), 400

        # Batch info
        batch_code = f"BATCH-{recipe_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        flash(f"Recipe batch started successfully!", "success")

    except pyodbc.Error as e:
        conn.rollback()
        flash(f"Database error: {str(e)}", "danger")
        return jsonify({"error": f"Database error: {str(e)}"}), 500
    except Exception as e:
        flash(f"Unexpected error: {str(e)}", "danger")
        return jsonify({"error": f"Unexpected error: {str(e)}"}), 500
    finally:
        conn.close()

    return jsonify(
        {"message": "Recipe batch started successfully", "batch_code": batch_code}
    )


def row_to_dict(cursor, row):
    columns = [col[0] for col in cursor.description]
    return {col: row[i] if i < len(row) else None for i, col in enumerate(columns)}


@app.route("/addTag", methods=["POST"])
def add_tag():
    """Add a new tag to the database"""

    # Parse JSON Request
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "error": "Invalid JSON payload"}), 400

    tag_id = data.get("tagId")
    tag_name = data.get("tagName")
    tag_address = data.get("tagAddress")
    plc_id = data.get("plcId")

    # Validate Required Fields
    if not all([tag_id, tag_name, tag_address, plc_id]):
        return jsonify({"success": False, "error": "Missing required fields."}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "error": "Database connection failed."}), 500

    try:
        with conn.cursor() as cursor:
            # Insert the new tag
            cursor.execute(
                """
                INSERT INTO Tag_Table (tagId, tagName, tagAddress, plcId)
                VALUES (?, ?, ?, ?)
                """,
                (tag_id, tag_name, tag_address, plc_id),
            )
            conn.commit()

        # Fetch and update live value for the new tag
        fetch_and_update_live_value(tag_address)

        return jsonify({"success": True, "message": "Tag added successfully."}), 201

    except pyodbc.IntegrityError:
        return jsonify({"success": False, "error": "TagId already exists."}), 409
    except pyodbc.Error as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        conn.close()


def fetch_and_update_live_value(tag_address):
    """Fetch live value for a tag and update the Live_Tags table."""

    # Establish database connection
    conn = get_db_connection()
    if not conn:
        print("Failed to connect to database.")
        return

    client = Client(ENDPOINT_URL)
    client.session_timeout = 30000  # Adjust timeout as needed

    try:
        client.connect()
        node = client.get_node(tag_address)
        value = node.get_value()

        # Convert the value into a JSON string if necessary
        if isinstance(value, (list, dict)):
            value = json.dumps(value)
        else:
            value = str(value)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with conn.cursor() as cursor:
            # Fetch tagId from Tag_Table
            cursor.execute(
                "SELECT Tagid FROM Tag_Table WHERE tagAddress = ?", (tag_address,)
            )
            tag_id = cursor.fetchone()

            if tag_id:
                tag_id = tag_id[0]

                # Use MERGE statement for upsert (INSERT or UPDATE)
                cursor.execute(
                    """
                    MERGE INTO Live_Tags AS target
                    USING (SELECT ? AS tagId) AS source
                    ON target.tagId = source.tagId
                    WHEN MATCHED THEN
                        UPDATE SET value = ?, timestamp = ?
                    WHEN NOT MATCHED THEN
                        INSERT (value, tagId, timestamp) VALUES (?, ?, ?);
                """,
                    (tag_id, value, timestamp, value, tag_id, timestamp),
                )

                conn.commit()

                # Emit live data to the frontend via SocketIO
                # socketio.emit('liveData', {"success": True, "tagAddress": tag_address, "value": value})

    except Exception as e:
        # socketio.emit('liveData', {"success": False, "error": str(e)})
        print(f"Error fetching and updating live value: {e}")

    finally:
        client.disconnect()
        conn.close()


def fetch_live_tag_data():
    """
    Fetch live tags and their corresponding tag addresses from the database.
    Returns a list of dictionaries.
    """
    conn = get_db_connection()
    if not conn:
        return []

    try:
        with conn.cursor() as cursor:
            query = """
                SELECT lt.id, lt.tagId, tt.tagAddress 
                FROM Live_Tags lt
                LEFT JOIN Tag_Table tt ON lt.tagId = tt.Tagid
            """
            cursor.execute(query)

            # Fetch results and return as a list of dictionaries
            columns = [column[0] for column in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    except pyodbc.Error as e:
        print(f"Error fetching live tag data: {e}")
        return []

    finally:
        conn.close()


def read_opcua_values(tag_data):
    """
    Connect to OPC UA server and read values for the given tag addresses.
    """
    client = Client(ENDPOINT_URL)
    client.session_timeout = 30000  # Adjust timeout as needed
    updates = []

    try:
        # print("Connecting to OPC UA server...")
        client.connect()
        # print("Connected to OPC UA server!")

        # Start the overall timing for the entire operation
        overall_start_time = time.time()

        for row in tag_data:
            # print("Processing row:", row)  # Debugging output

            # Ensure the dictionary contains required keys
            if not isinstance(row, dict) or not all(
                k in row for k in ("id", "tagId", "tagAddress")
            ):
                # print(f"Invalid row format: {row}, skipping...")
                continue

            live_tag_id = row["id"]  # Extract 'id'
            tag_id = row["tagId"]  # Extract 'tagId'
            tag_address = row["tagAddress"]  # Extract 'tagAddress'

            # print(f"Processing tagId {tag_id} with tagAddress: {tag_address}")

            if not tag_address:
                print(f"No tagAddress found for tagId {tag_id}. Skipping update.")
                continue

            # Start timing for individual tag fetching
            tag_start_time = time.time()

            try:
                # Fetch the live value from OPC UA server
                node = client.get_node(tag_address)
                value = node.get_value()

                # Calculate individual tag fetching time
                tag_end_time = time.time()
                tag_elapsed_time = (
                    tag_end_time - tag_start_time
                ) * 1000  # in milliseconds
                # print(f"Fetched value {value} for tagId {tag_id}. Time taken: {tag_elapsed_time:.2f} ms.")

                # Store the result
                updates.append((value, live_tag_id))
            except Exception as e:
                print(f"Error fetching live value for tagId {tag_id}: {e}")

        # Calculate overall session duration
        overall_end_time = time.time()
        overall_elapsed_time = (
            overall_end_time - overall_start_time
        ) * 1000  # in milliseconds
        # print(f"Disconnected from OPC UA server. Overall session duration: {overall_elapsed_time:.2f} ms.")

    finally:
        client.disconnect()

    return updates


def update_database(updates):
    """
    Bulk update the Live_Tags table with new live values.
    Converts lists or dictionaries in 'value' to JSON strings.
    """
    if not updates:
        return

    updates_serialized = []  # Store serialized updates for Live_Tags
    consistent_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for value, record_id in updates:
        if isinstance(value, (list, dict)):
            value = json.dumps(value)  # Convert list/dict to JSON string
        else:
            value = str(value)  # Convert other types to string

        # Prepare data for bulk update
        updates_serialized.append((value, consistent_timestamp, record_id))

    # Database operations
    conn = get_db_connection()
    if not conn:
        return

    try:
        with conn.cursor() as cursor:
            cursor.executemany(
                """
                UPDATE Live_Tags
                SET value = ?, timestamp = ?
                WHERE id = ?
            """,
                updates_serialized,
            )

            conn.commit()
            # print(f"✅ {len(updates_serialized)} records updated successfully!")

    except pyodbc.Error as e:
        print(f"Error updating Live_Tags: {e}")

    finally:
        conn.close()


# API_URL = "https://api.ngsmart.in:5000/suvi/api/v1/machine-data"  # Change this to your Flask API URL
# API_URL = "https://api.ngsmart.in:5000/suvi/api/v1/machine-data"
API_URL = "http://localhost:4000/suvi/api/v1/machine-data"
BATCH_SIZE = 10


def is_json(value):
    """
    Check if a value is a valid JSON.
    """
    try:
        json.loads(value)
        return True
    except ValueError:
        return False


def process_value(tag_name, value):
    """
    Process the value field:
    - If it's an array, convert each element into a separate field like field1, field2, etc.
    - If it's not an array, return as-is.
    """
    if isinstance(value, list):
        return {f"{tag_name}{i+1}": v for i, v in enumerate(value)}
    else:
        return {tag_name: value}


def timestamp_to_epoch(timestamp):
    """
    Convert timestamp from 'YYYY-MM-DD HH:MM:SS' (string) or datetime.datetime object to epoch seconds.
    If the timestamp is already in epoch form (integer), return it directly.
    """
    try:
        # If timestamp is an integer (epoch time)
        if isinstance(timestamp, int):
            return timestamp

        # If timestamp is a string in 'YYYY-MM-DD HH:MM:SS' format
        elif isinstance(timestamp, str):
            # Handle milliseconds if present
            if "." in timestamp:
                datetime_obj = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S.%f")
            else:
                datetime_obj = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
            return int(datetime_obj.timestamp())

        # If timestamp is a datetime.datetime object
        elif isinstance(timestamp, datetime):
            return int(timestamp.timestamp())

        # Handle unexpected timestamp types
        else:
            print(f"Unexpected timestamp type: {type(timestamp)}")
            return None
    except ValueError as e:
        print(f"Error converting timestamp: {e}")
        return None


def get_current_batch_count():
    """
    Fetch the current total number of batches from the database.
    """
    conn = get_db_connection()
    if not conn:
        return 0  # Return 0 if unable to connect to DB

    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT number_of_batches_created FROM BatchTracker WHERE id = 1"
            )
            result = cursor.fetchone()
            return int(result[0]) if result and result[0] is not None else 0

    except pyodbc.Error as e:
        print(f"Error fetching batch count: {e}")
        return 0

    finally:
        conn.close()


def update_batch_count(new_batches):
    """
    Add the count of new batches to the existing total and update the database.
    """
    conn = get_db_connection()
    if not conn:
        return None  # Return None if DB connection fails

    try:
        with conn.cursor() as cursor:
            # Fetch the current batch count
            cursor.execute(
                "SELECT number_of_batches_created FROM BatchTracker WHERE id = 1"
            )
            result = cursor.fetchone()
            current_count = int(result[0]) if result and result[0] is not None else 0

            # print(f"Current batch count: {current_count}")

            # Calculate updated count
            updated_count = current_count + new_batches

            # Update the database with the new count
            cursor.execute(
                "UPDATE BatchTracker SET number_of_batches_created = ? WHERE id = 1",
                (updated_count,),
            )
            conn.commit()

            # print(f"Updated batch count: {updated_count}")
            return updated_count

    except pyodbc.Error as e:
        print(f"Error updating batch count: {e}")
        return None

    finally:
        conn.close()


def clean_live_log_last_batch(record_ids, chunk_size=500):
    """
    Clean up successfully uploaded records from MSSQL in chunks to avoid parameter limit errors.
    """
    conn = get_db_connection()
    if not conn:
        return False  # Return False if DB connection fails

    try:
        cursor = conn.cursor()

        # Process deletions in chunks
        for i in range(0, len(record_ids), chunk_size):
            chunk = record_ids[i : i + chunk_size]

            # Create dynamic placeholders
            placeholders = ", ".join(["?"] * len(chunk))
            query = f"DELETE FROM Live_Log WHERE id IN ({placeholders})"

            # Execute batch delete
            cursor.execute(query, chunk)
            conn.commit()

        print("✅ Successfully cleaned up uploaded records from MSSQL.")
        return True

    except pyodbc.Error as e:
        print(f"❌ Error during cleanup: {e}")
        return False

    finally:
        conn.close()


def fetch_plc_info():
    """Fetch plcId and serialKey from Plc_Table in MSSQL."""
    conn = get_db_connection()
    if not conn:
        return None, None  # Return None if DB connection fails

    try:
        cursor = conn.cursor()

        # Use "TOP 1" instead of "LIMIT 1" for MSSQL
        cursor.execute("SELECT TOP 1 plcId, serial_Key FROM Plc_Table")
        plc_info = cursor.fetchone()

        return plc_info if plc_info else (None, None)

    except pyodbc.Error as e:
        print(f"❌ Error fetching PLC info: {e}")
        return None, None

    finally:
        conn.close()  # Ensure the connection is closed


def upload_live_log_to_mongodb():
    """
    Fetch records from MSSQL, format them, send them to Flask API, and clean up MSSQL.
    """
    try:
        # Fetch plcId and serialKey from Plc_Table
        plc_id, serial_key = fetch_plc_info()
        if plc_id is None or serial_key is None:
            print("❌ Error: plcId or serialKey not found in Plc_Table.")
            return

        # Connect to MSSQL
        conn = get_db_connection()
        if not conn:
            return

        try:
            cursor = conn.cursor()

            # Fetch all records grouped by timestamp
            cursor.execute('''
                SELECT id, tagName, value, timestamp
                FROM Live_Log
                ORDER BY timestamp ASC, id ASC
            ''')
            records = cursor.fetchall()

            if not records:
                print("✅ No records to upload.")
                return

            # Group records by timestamp
            grouped_data = {}
            record_ids = []

            for record in records:
                record_id, tag_name, value, timestamp = record

                # Convert timestamp to epoch format
                epoch_timestamp = timestamp_to_epoch(timestamp)
                if epoch_timestamp is None:
                    continue  # Skip if timestamp conversion failed

                # Parse value if it's JSON
                parsed_value = json.loads(value) if is_json(value) else value

                # Process value to split arrays into separate fields
                processed_data = process_value(tag_name, parsed_value)

                if epoch_timestamp not in grouped_data:
                    grouped_data[epoch_timestamp] = []
                grouped_data[epoch_timestamp].append(processed_data)

                # Collect record IDs for cleanup
                record_ids.append(record_id)

            # Prepare the batch format
            batches = []
            for epoch_timestamp, tags in grouped_data.items():
                # Combine all tags for the same timestamp into one dictionary
                combined_tags = {}
                for tag in tags:
                    combined_tags.update(tag)

                batch = {
                    "plcId": plc_id,
                    "serialNo": serial_key,
                    "values": [{
                        "dbId": 3,
                        "dbNo": 1001,
                        "dbName": "dbCloud",
                        "data": [{
                            "temp": 1,
                            "timeStamp": epoch_timestamp * 1000,
                            **combined_tags
                        }]
                    }]
                }
                batches.append(batch)

            successful_batches = 0

            # Send data in batches of 10
            for i in range(0, len(batches), BATCH_SIZE):
                batch_to_send = batches[i:i + BATCH_SIZE]

                response = requests.post(API_URL, json=batch_to_send)

                if response.status_code in [200, 201]:
                    print(f"✅ Batch {i // BATCH_SIZE + 1} uploaded successfully.")
                    successful_batches += 1

                    # Clean up uploaded records
                    clean_live_log_last_batch(record_ids, chunk_size=500)
                else:
                    print(f"❌ Failed to upload batch {i // BATCH_SIZE + 1}. Status code: {response.status_code}")
                    print(f"Response: {response.text}")

            # Update batch count in MSSQL
            update_batch_count(successful_batches)
            print("✅ Process completed.")

        except pyodbc.Error as e:
            print(f"❌ Database query error: {e}")

        finally:
            conn.close()  # Ensure the connection is closed

    except Exception as e:
        print(f"❌ Error: {e}")

def schedule_live_log_upload_background(interval=100000):
    """
    Run the upload function in a background thread every `interval` milliseconds.
    """
    def background_task():
        while True:
            try:
                conn = get_db_connection()
                if not conn:
                    print("❌ Skipping task due to DB connection failure.")
                    time.sleep(interval / 1000)
                    continue

                try:
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM Live_Log")
                    count = cursor.fetchone()[0]

                    if count > 1000:
                        print("📢 More than 1000 records found. Uploading...")
                        upload_live_log_to_mongodb()
                    else:
                        print("ℹ️ Less than 1000 records in Live_Log. Skipping upload...")

                except pyodbc.Error as e:
                    print(f"❌ Error executing query: {e}")

                finally:
                    conn.close()  # Ensure connection is closed properly

            except Exception as e:
                print(f"❌ Error in background task: {e}")

            time.sleep(interval / 1000)  # Wait before the next check

    # Start the background task in a separate daemon thread
    # thread = threading.Thread(target=background_task, daemon=True)
    # thread.start()

# Call this function to start the background thread
# schedule_live_log_upload_background(interval=15000)


def update_all_live_tags():
    """
    Fetch all entries from Live_Tags, resolve tagAddress from Tag_Table, and update live values.
    """
    try:
        # Step 1: Fetch data from the database
        tag_data = fetch_live_tag_data()

        # Step 2: Read values from OPC UA
        updates = read_opcua_values(tag_data)

        # Step 3: Update the database
        update_database(updates)

    except Exception as e:
        print(f"Error: {e}")
        # Global variable to store the user-defined interval (in seconds)


def update_all_live_tags_to_log():
    """
    Insert values into Live_Log table based on user-defined interval.
    """
    try:
        # Fetch the latest live tag data
        tag_data = fetch_live_tag_data()

        # Read values from OPC UA
        updates = read_opcua_values(tag_data)

        # Generate a single consistent timestamp for all records
        consistent_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Prepare live log entries
        live_log_entries = []
        for value, record_id in updates:
            if isinstance(value, (list, dict)):
                value = json.dumps(value)  # Convert list/dict to JSON string
            else:
                value = str(value)  # Convert other types to string
            live_log_entries.append((value, consistent_timestamp, record_id))

        # Insert into Live_Log table
        conn = get_db_connection()
        if not conn:
            print("❌ Database connection failed. Skipping update.")
            return

        try:
            cursor = conn.cursor()

            sql_query = """
                INSERT INTO Live_Log (tagId, tagName, value, timestamp)
                SELECT lt.tagId, tt.tagName, ?, ?
                FROM Live_Tags lt
                LEFT JOIN Tag_Table tt ON lt.tagId = tt.Tagid
                WHERE lt.id = ?
            """

            cursor.executemany(
                sql_query, live_log_entries
            )  # Batch insert for efficiency
            conn.commit()
            print(
                f"✅ {len(live_log_entries)} records inserted into Live_Log successfully."
            )

        except pyodbc.Error as e:
            print(f"❌ SQL Execution Error: {e}")

        finally:
            conn.close()  # Ensure connection is closed properly

    except Exception as e:
        print(f"❌ Error in update_all_live_tags_to_log: {e}")


@app.route("/writeValue", methods=["POST"])
def write_value():
    data = request.json
    print("Received Data:", data)  # This will print the data to the console
    node_id = data.get("nodeId")
    value = data.get("value")

    if not node_id or value is None:
        return jsonify(
            {"success": False, "error": "Missing nodeId or value in the request."}
        )

    try:
        # Connect to the OPC UA server
        client = Client(ENDPOINT_URL)
        client.session_timeout = 30000  # Adjust timeout as needed
        client.connect()

        # Get the node object 'ns=3;db.recipe

        node = client.get_node(node_id)
        print("Received node:", node)  # This will print the data to the console

        # Fetch the DataType of the node
        data_type = node.get_data_type_as_variant_type()
        # print("Received datatype:", data_type)  # This will print the data to the console
        if not data_type:
            raise Exception(f"Could not fetch DataType for NodeId: {node_id}")

        # print(f"Fetched DataType: {data_type}")

        # Wrap the value in a ua.Variant object with the correct DataType
        # Wrap the value in a ua.Variant object with the correct DataType
        variant = ua.DataValue(ua.Variant(value, data_type))

        # Perform the write operation
        node.set_value(variant)

        # Disconnect from the client
        client.disconnect()

        return jsonify(
            {
                "success": True,
                "message": f"Value written successfully to NodeId: {node_id}",
            }
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/readValues", methods=["GET"])
def read_values():
    try:
        # Connect to SQL Database
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "error": "Database connection failed"})

        cursor = conn.cursor()

        # Fetch node IDs and their corresponding tag names
        cursor.execute(
            "SELECT DISTINCT tagAddress, tagName FROM Tag_Table"
        )  # Include tagName
        tag_data = cursor.fetchall()

        if not tag_data:
            conn.close()
            return jsonify(
                {"success": False, "error": "No tag data found in the database"}
            )

        # Connect to the PLC
        client = Client(ENDPOINT_URL)
        client.session_timeout = 30000  # Adjust timeout as needed
        client.connect()

        results = []
        for tag_address, tag_name in tag_data:
            try:
                node = client.get_node(tag_address)
                value = node.get_value()
                timestamp = datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                )  # Add current timestamp
                results.append(
                    {
                        "nodeId": tag_address,
                        "tagName": tag_name,
                        "value": value,
                        "timestamp": timestamp,
                    }
                )
            except Exception as e:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                results.append(
                    {
                        "nodeId": tag_address,
                        "tagName": tag_name,
                        "error": str(e),
                        "timestamp": timestamp,
                    }
                )

        # Disconnect client & close DB connection
        client.disconnect()
        conn.close()

        return jsonify({"success": True, "results": results})

    except pyodbc.Error as db_error:
        return jsonify({"success": False, "error": f"Database error: {str(db_error)}"})

    except Exception as e:
        return jsonify({"success": False, "error": f"Unexpected error: {str(e)}"})


def read_live_values(node_ids):
    """Connect to OPC UA server and emit live data to the frontend."""
    client = Client(ENDPOINT_URL)
    client.session_timeout = 30000  # Adjust timeout as needed
    try:
        # print("Attempting to connect to OPC UA server...")
        client.connect()
        # print("Connected to OPC UA server!")

        while True:
            results = []
            # print("Fetching data for node IDs:", node_ids)  # Debug node IDs
            for node_id in node_ids:
                try:
                    # print(f"Reading value for node ID: {node_id}")  # Log node ID
                    node = client.get_node(node_id)
                    value = node.get_value()
                    # print(f"Value for node {node_id}: {value}")  # Log fetched value
                    results.append({"nodeId": node_id, "value": value})
                except Exception as e:
                    print(f"Error reading node {node_id}: {e}")  # Log errors
                    results.append({"nodeId": node_id, "error": str(e)})

            # print("Emitting data to frontend:", results)  # Debug emitted data
            # socketio.emit('liveData', {"success": True, "results": results})
            time.sleep(1)  # Fetch data every 2 seconds
    except Exception as e:
        print(f"Error in OPC UA connection: {e}")  # Log connection errors
        # socketio.emit('liveData', {"success": False, "error": str(e)})
    finally:
        client.disconnect()
        # print("Disconnected from OPC UA server!")  # Debug disconnection


@app.route("/getLiveValues", methods=["GET"])
def get_live_values():
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "error": "Database connection failed"})

        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT Live_Tags.Id, Live_Tags.value, Live_Tags.timestamp, Tag_Table.tagName
            FROM Live_Tags
            INNER JOIN Tag_Table ON Live_Tags.tagId = Tag_Table.tagId
        """
        )

        live_values = cursor.fetchall()

        results = [
            {"tagName": row.tagName, "value": row.value, "timestamp": row.timestamp}
            for row in live_values
        ]

        conn.close()
        return jsonify({"success": True, "data": results})

    except pyodbc.Error as db_error:
        return jsonify({"success": False, "error": f"Database error: {str(db_error)}"})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/startLiveRead", methods=["POST"])
def start_live_read():
    """API endpoint to initiate live data reading."""
    data = request.json
    print("Received request to start live reading:", data)  # Log request data

    node_ids = data.get("nodeIds", [])
    if not node_ids:
        print("No node IDs provided in request!")  # Log missing node IDs
        return jsonify({"success": False, "message": "No node IDs provided"}), 400

    print("Starting live reading for node IDs:", node_ids)  # Log starting point
    # Start the live reading thread
    thread = threading.Thread(target=read_live_values, args=(node_ids,))
    thread.daemon = True  # Ensures thread stops when the main app stops
    thread.start()

    return jsonify({"success": True, "message": "Live data reading started"})


@app.route("/get_all_recipes", methods=["GET"])
def get_all_recipes():
    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"})

    try:
        cursor = conn.cursor()
        cursor.execute("SELECT Recipe_ID, Recipe_Name FROM Recipe")
        recipes = [
            {"recipe_id": row[0], "recipe_name": row[1]} for row in cursor.fetchall()
        ]

        return jsonify({"success": True, "recipes": recipes})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

    finally:
        conn.close()


@app.route("/get_recipe", methods=["GET"])
def get_recipe():
    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "message": "Database connection failed"})

    recipe_id = request.args.get("recipe_id")
    recipe_name = request.args.get("recipe_name")

    print("📌 Recipe ID received:", recipe_id, "Recipe Name:", recipe_name)

    if not recipe_id and not recipe_name:
        return jsonify(
            {"success": False, "message": "Recipe ID or Recipe Name is required"}
        )

    try:
        cursor = conn.cursor()

        # Fetch data from Recipe table
        if recipe_id:
            cursor.execute("SELECT * FROM Recipe WHERE recipe_id = ?", (recipe_id,))
        else:
            cursor.execute("SELECT * FROM Recipe WHERE recipe_name = ?", (recipe_name,))
        recipe_main = cursor.fetchone()

        if not recipe_main:
            return jsonify({"success": False, "message": "Recipe not found"})

        # Extract column names
        columns = [column[0] for column in cursor.description]
        recipe_main_dict = dict(zip(columns, recipe_main))

        # Fetch Recipe_Details1 (Spacer, Alu_roller_type, materials, positions)
        cursor.execute(
            "SELECT * FROM Recipe_Details1 WHERE recipe_id = ?",
            (recipe_main_dict["Recipe_ID"],),
        )
        recipe_pos = cursor.fetchall()
        recipe_pos_list = [
            {column[0]: value for column, value in zip(cursor.description, row)}
            for row in recipe_pos
        ]

        # Fetch Sub_Menu (machine settings)
        cursor.execute(
            "SELECT * FROM Sub_Menu WHERE recipe_id = ?",
            (recipe_main_dict["Recipe_ID"],),
        )
        recipe_motor = cursor.fetchone()
        recipe_motor_dict = (
            {
                column[0]: value
                for column, value in zip(cursor.description, recipe_motor)
            }
            if recipe_motor
            else {}
        )

        # Fetch Inspection_Settings
        cursor.execute(
            "SELECT * FROM Inspection_Settings WHERE Recipe_ID = ?",
            (recipe_main_dict["Recipe_ID"],),
        )
        inspection_data = cursor.fetchone()
        inspection_data_dict = (
            {
                column[0]: value
                for column, value in zip(cursor.description, inspection_data)
            }
            if inspection_data
            else {}
        )

        # Build unified response
        data = {
            "recipe_id": recipe_main_dict.get("Recipe_ID"),
            "recipe_name": recipe_main_dict.get("Recipe_Name"),
            "filter_size": recipe_main_dict.get("Filter_Size"),
            "filter_code": recipe_main_dict.get("Filter_Code"),
            "art_no": recipe_main_dict.get("Art_No"),
            # Machine settings
            "Pleat_Height": recipe_motor_dict.get("Pleat_Height", ""),
            "Blade_Opening": recipe_motor_dict.get("Blade_Opening", ""),
            "Soft_Touch": recipe_motor_dict.get("Soft_Touch", ""),
            "Press_Touch": recipe_motor_dict.get("Press_Touch", ""),
            "Feeder1_Media_Thickness": recipe_motor_dict.get(
                "Feeder1_Media_Thickness", ""
            ),
            "Feeder1_Park_Position": recipe_motor_dict.get("Feeder1_Park_Position", ""),
            "Feeder2_Media_Thickness": recipe_motor_dict.get(
                "Feeder2_Media_Thickness", ""
            ),
            "Feeder2_Park_Position": recipe_motor_dict.get("Feeder2_Park_Position", ""),
            "Foil1_Length": recipe_motor_dict.get("Foil1_Length", ""),
            "Foil1_Length_Offset": recipe_motor_dict.get("Foil1_Length_Offset", ""),
            "Foil1_Width": recipe_motor_dict.get("Foil1_Width", ""),
            "Foil1_Tension_Set_Point": recipe_motor_dict.get(
                "Foil1_Tension_Set_Point", ""
            ),
            "Foil2_Length": recipe_motor_dict.get("Foil2_Length", ""),
            "Foil2_Length_Offset": recipe_motor_dict.get("Foil2_Length_Offset", ""),
            "Foil2_Width": recipe_motor_dict.get("Foil2_Width", ""),
            "Foil2_Tension_Set_Point": recipe_motor_dict.get(
                "Foil2_Tension_Set_Point", ""
            ),
            "Puller_Start_Position": recipe_motor_dict.get("Puller_Start_Position", ""),
            "Puller_End_Position": recipe_motor_dict.get("Puller_End_Position", ""),
            "Puller2_Feed_Correction": recipe_motor_dict.get(
                "Puller2_Feed_Correction", ""
            ),
            "Puller_Extra_Stroke_Enable": recipe_motor_dict.get(
                "Puller_Extra_Stroke_Enable", ""
            ),
            "Autofeeding_Offset": recipe_motor_dict.get("Autofeeding_Offset", ""),
            "Sync_Table_Start_Position": recipe_motor_dict.get(
                "Sync_Table_Start_Position", ""
            ),
            "Pack_Transfer_Rev_Position": recipe_motor_dict.get(
                "Pack_Transfer_Rev_Position", ""
            ),
            "Pack_Transfer_Park_Position": recipe_motor_dict.get(
                "Pack_Transfer_Park_Position", ""
            ),
            "Filter_Box_Height": recipe_motor_dict.get("Filter_Box_Height", ""),
            "Filter_Box_Width": recipe_motor_dict.get("Filter_Box_Width", ""),
            "Filter_Box_Length": recipe_motor_dict.get("Filter_Box_Length", ""),
            "Pleat_Pitch": recipe_motor_dict.get("Pleat_Pitch", ""),
            "Pleat_Counts": recipe_motor_dict.get("Pleat_Counts", ""),
            "Batch_Count": recipe_motor_dict.get("Batch_Count", ""),
            "Foil_Low_Diameter": recipe_motor_dict.get("Foil_Low_Diameter", ""),
            "Media_Tension_Set_Point": recipe_motor_dict.get(
                "Media_Tension_Set_Point", ""
            ),
            "Feeding_Conveyor_Speed": recipe_motor_dict.get(
                "Feeding_Conveyor_Speed", ""
            ),
            "Discharge_Conveyor_Speed": recipe_motor_dict.get(
                "Discharge_Conveyor_Speed", ""
            ),
            "Lid_Placement_Enable": recipe_motor_dict.get("Lid_Placement_Enable", ""),
            "Lid_Placement_Position": recipe_motor_dict.get(
                "Lid_Placement_Position", ""
            ),
            # Air Inspection Settings
            "databaseAvailable": inspection_data_dict.get("databaseAvailable", ""),
            "Width": inspection_data_dict.get("Width", ""),
            "Height": inspection_data_dict.get("Height", ""),
            "Depth": inspection_data_dict.get("Depth", ""),
            "Art_Num": inspection_data_dict.get("Art_No", ""),
            "Air_Flow_Set": inspection_data_dict.get("Air_Flow_Set", ""),
            "Pressure_Drop_Setpoint": inspection_data_dict.get(
                "Pressure_Drop_Setpoint", ""
            ),
            "Lower_Tolerance1": inspection_data_dict.get("Lower_Tolerance1", ""),
            "Lower_Tolerance2": inspection_data_dict.get("Lower_Tolerance2", ""),
            "Upper_Tolerance1": inspection_data_dict.get("Upper_Tolerance1", ""),
            "Upper_Tolerance2": inspection_data_dict.get("Upper_Tolerance2", ""),
        }

        # Overwrite materials + positions from Recipe_Details1
        for pos in recipe_pos_list:
            data["Alu_roller_type"] = pos.get("Alu_roller_type", "")
            data["Spacer"] = pos.get("Spacer", "")
            data["Alu_Material"] = pos.get("Alu_Material", "")
            data["House_Material"] = pos.get("House_Material", "")
            for i in range(1, 10):
                pos_key = f"Pos{i}"
                if pos_key in pos:
                    data[pos_key] = pos[pos_key]

        print("📌 Final response data:", data)
        return jsonify({"success": True, **data})

    except pyodbc.Error as db_error:
        print(f"❌ Database error: {db_error}")
        return jsonify(
            {"success": False, "message": f"Database error: {str(db_error)}"}
        )

    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        return jsonify({"success": False, "message": str(e)})

    finally:
        conn.close()


@app.route("/report", methods=["GET", "POST"])
def report():
    conn = get_db_connection()
    cursor = conn.cursor()

    # --- Pagination setup ---
    page = request.args.get("page", 1, type=int)
    per_page = 10
    offset = (page - 1) * per_page

    # --- Date filters ---
    start_date = request.values.get("start_date")
    end_date = request.values.get("end_date")

    base_query = "FROM Recipe_Log"
    where_clauses = []
    params = []

    # --- Fetch column names dynamically once ---
    cursor.execute(
        "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'Recipe_Log'"
    )
    columns = [col[0] for col in cursor.fetchall()]

    # --- Column search filters ---
    column_filters = []
    filter_params = []
    for col_name in columns:
        value = request.values.get(col_name)
        if value:
            column_filters.append(f"[{col_name}] LIKE ?")
            filter_params.append(f"%{value}%")

    # --- Combine all filters ---
    if start_date and end_date:
        where_clauses.append("CONVERT(DATE, Timestamp) BETWEEN ? AND ?")
        params.extend([start_date, end_date])

    if column_filters:
        where_clauses.extend(column_filters)
        params.extend(filter_params)

    if where_clauses:
        base_query += " WHERE " + " AND ".join(where_clauses)

    # --- Fetch paginated data ---
    query = f"SELECT * {base_query} ORDER BY Timestamp DESC OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
    cursor.execute(query, params + [offset, per_page])
    rows = cursor.fetchall()

    # --- Convert rows to dict ---
    data = [dict(zip(columns, row)) for row in rows]

    # --- Count total records ---
    count_query = f"SELECT COUNT(*) {base_query}"
    cursor.execute(count_query, params)
    total_count = cursor.fetchone()[0]
    total_pages = (total_count + per_page - 1) // per_page

    start_entry = offset + 1 if total_count > 0 else 0
    end_entry = min(offset + per_page, total_count)

    conn.close()

    return render_template(
        "report.html",
        columns=columns,
        data=data,
        page=page,
        total_pages=total_pages,
        total_count=total_count,
        start_entry=start_entry,
        end_entry=end_entry,
        start_date=start_date,
        end_date=end_date,
    )


@app.route("/search-report")
def search_report():
    query = request.args.get("query", "").lower()
    conn = get_db_connection()
    cursor = conn.cursor()

    # Search in all relevant columns
    sql_query = f"""
        SELECT * FROM Recipe_Log 
        WHERE 
            LOWER(Batch_Code) LIKE ? OR
            LOWER(Timestamp) LIKE ? OR
            LOWER(Recipe_Name) LIKE ? OR
            LOWER(Art_No) LIKE ? OR
            LOWER(Filter_Size) LIKE ? OR
            LOWER(FilterSize) LIKE ? OR
            LOWER(NgStatus) LIKE ? OR
            LOWER(SerialNo) LIKE ? OR
            LOWER(Avg_Air_Flow) LIKE ? OR
            LOWER(Avg_Result) LIKE ? OR
            LOWER(Batch_Completion_Status) LIKE ?
    """

    cursor.execute(sql_query, tuple(f"%{query}%" for _ in range(11)))
    columns = [column[0] for column in cursor.description]
    data = [dict(zip(columns, row)) for row in cursor.fetchall()]

    conn.close()
    return jsonify(data)


@app.route("/download_excel", methods=["GET"])
def download_excel():
    conn = get_db_connection()
    cursor = conn.cursor()

    # --- Get filters ---
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # --- Base query ---
    base_query = "FROM Recipe_Log"
    where_clauses = []
    params = []

    # --- Get columns dynamically ---
    cursor.execute(
        "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'Recipe_Log'"
    )
    columns = [col[0] for col in cursor.fetchall()]

    # --- Column filters ---
    column_filters = []
    filter_params = []
    for col_name in columns:
        value = request.args.get(col_name)
        if value:
            column_filters.append(f"[{col_name}] LIKE ?")
            filter_params.append(f"%{value}%")

    # --- Apply date and column filters ---
    if start_date and end_date:
        where_clauses.append("CONVERT(DATE, Timestamp) BETWEEN ? AND ?")
        params.extend([start_date, end_date])
    if column_filters:
        where_clauses.extend(column_filters)
        params.extend(filter_params)

    if where_clauses:
        base_query += " WHERE " + " AND ".join(where_clauses)

    # --- Fetch all filtered data ---
    query = f"SELECT * {base_query} ORDER BY Timestamp DESC"
    cursor.execute(query, params)
    rows = cursor.fetchall()

    # --- Create Excel workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered Data"

    # Header row
    ws.append(columns)

    # Data rows
    for row in rows:
        ws.append(list(row))

    # --- Save to memory ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    conn.close()

    # --- Return file ---
    return send_file(
        output,
        as_attachment=True,
        download_name="Filtered_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


frontend_to_plc_map = {
    # Step 3: Machine Settings
    "Pleat_Height": "Pleat_Height",
    "Soft_Touch": "Soft_Touch",
    "Feeder1_Media_Thickness": "Left_Blade_MediaTHickness",
    "Feeder1_Park_Position": "Left_Blade_Start_pos",
    "Foil1_Length": "Corr1Feed_length",
    "Foil1_Length_Offset": "Corr1feed_length_offset",
    "Foil1_Width": "Corr1feed_length_width",
    "Puller_Start_Position": "Puller_Start_pos",
    "Autofeeding_Offset": "Auto_Feeding_Offset",
    "Filter_Box_Height": "Filter_Box_Height",
    "Filter_Box_Length": "Filter_Box_Length",
    "Pleat_Pitch": "Set_Pleat_Pitch",
    "Pleat_Counts": "Set_Pleat_Count",
    "Foil_Low_Diameter": "Low_Dia_Set",
    "Feeding_Conveyor_Speed": "Feeding_Conveyor_Speed",
    "Pack_Transfer_Rev_Position": "Filter_Transfer_pos_rev",
    "Foil1_Tension_Set_Point": "Corr_1_Set_point",
    "Foil2_Tension_Set_Point": "Corr_2_Set_point",
    "Blade_Opening": "Blade_opening",
    "Press_Touch": "Press_Touch",
    "Feeder2_Media_Thickness": "Right_Blade_MediaThickness",
    "Feeder2_Park_Position": "Right_Blade_Start_pos",
    "Foil2_Length": "Corr2feed_length",
    "Foil2_Length_Offset": "Corr2feed_length_offset",
    "Foil2_Width": "Corr2feed_length_width",
    "Puller_End_Position": "Puller_End_pos",
    "Puller2_Feed_Correction": "Puller2_Feed_Correction",
    "Puller_Extra_Stroke_Enable": "PULLER_Extra stoke_Enable",
    "Filter_Box_Width": "Filter_Box_Width",
    "Lid_Placement_Enable": "Lid_Placement_Enable",
    "Lid_Placement_Position": "Lid_Placement_Pos",
    "Sync_Table_Start_Position": "Sync_Table_Start_Counts",
    "Batch_Count": "Set_Batch_COunt",
    "Discharge_Conveyor_Speed": "Discharge_COnvyor_Speed",
    "Pack_Transfer_Park_Position": "Filter_Transfer_pos_fwd",
    "Media_Tension_Set_Point": "Media_puller_2_SP",
}
frontend_to_plc_map2 = {
    "databaseAvailable": "databaseAvailable",
    "Width": "Width",
    "Height": "Height",
    "Depth": "Depth",
    "Art_Num": "Art No.",  # Mapping frontend name to OPC node
    "Air_Flow_Set": "Air Flow set",  # Adjusted mapping
    "Pressure_Drop_Setpoint": "Pressure Drop Setpoint",
    "Lower_Tolerance1": "Lower_Tolerance1",
    "Lower_Tolerance2": "Lower_Tolerance2",
    "Upper_Tolerance1": "Upper_Tolerance1",
    "Upper_Tolerance2": "Upper_Tolerance2",
    # "Lower_Fan_Speed": "Lower_fan_speed",
    # "Upper_Fan_Speed": "Upper_fan_Speed"
}


@app.route("/get_last_plc_values", methods=["GET"])
def get_last_plc_values():
    from opcua import Client

    client = Client(ENDPOINT_URL)
    try:
        client.connect()
        print("Connected to OPC UA Server")

        base_node = 'ns=3;s="dbRecipe1"."Recipe".'

        last_plc_values = {}

        for frontend_field, plc_field in frontend_to_plc_map.items():
            node_id = f'{base_node}"{plc_field}"'
            try:
                node = client.get_node(node_id)
                value = node.get_value()
                last_plc_values[frontend_field] = str(value)
            except Exception as e:
                print(f"Error fetching {frontend_field} ({node_id}): {e}")
                last_plc_values[frontend_field] = None

        for frontend_field, plc_field in frontend_to_plc_map2.items():
            node_id = f'{Test_Complete1}"{plc_field}"'
            try:
                node = client.get_node(node_id)
                value = node.get_value()
                last_plc_values[frontend_field] = str(value)
            except Exception as e:
                print(f"Error fetching Air field {frontend_field} ({node_id}): {e}")
                last_plc_values[frontend_field] = None

        return jsonify(last_plc_values)

    except Exception as e:
        print("Error fetching PLC values:", e)
        return jsonify({"error": "Failed to fetch PLC values"}), 500
    finally:
        client.disconnect()
        print("Disconnected from OPC UA Server")


def run_periodic_update1():
    while True:

        update_all_live_tags()

        time.sleep(1)  # Update every 1 seconds


# Initialize the default update interval
update_interval = 20
interval_lock = threading.Lock()


@app.route("/api/update_interval_time", methods=["POST"])
def update_interval_time():
    global update_interval

    try:
        # Get the intervalTime value from the request body
        data = request.json
        interval_time = data.get("intervalTime")

        if not interval_time or interval_time <= 0:
            return jsonify(
                {"success": False, "error": "Invalid interval time provided."}
            )

        # Update the intervalTime value in the MSSQL database
        conn = get_db_connection()
        if not conn:
            return jsonify({"success": False, "error": "Database connection failed."})

        try:
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE Plc_Table SET intervalTime = ? WHERE plcId = 1",
                (interval_time,),
            )
            conn.commit()
        except pyodbc.Error as db_error:
            print(f"❌ Database error: {db_error}")
            return jsonify({"success": False, "error": "Database update failed."})
        finally:
            conn.close()

        # Safely update the global update_interval value
        with interval_lock:
            update_interval = interval_time

        return jsonify(
            {"success": True, "message": "Interval time updated successfully."}
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def run_periodic_update3():
    global update_interval

    while True:
        # Perform the periodic update task
        update_all_live_tags_to_log()  # Insert data into Live_Log after the interval

        # Safely access the update_interval value
        with interval_lock:
            # print("Interval time:",update_interval)
            sleep_time = update_interval

        time.sleep(sleep_time)  # Sleep for the specified interval


if __name__ == "__main__":
    print("Starting Flask app...")
    # schedule_live_log_upload_background(interval=10000)
    # threading.Thread(target=run_periodic_update1, daemon=True).start()
    # threading.Thread(target=run_periodic_update3, daemon=True).start()
    thread = threading.Thread(target=opcua_monitoring, daemon=True).start()
    thread = threading.Thread(target=opcua_qrcode_monitoring, daemon=True).start()
    thread = threading.Thread(target=update_batch_status, daemon=True).start()
    thread=threading.Thread(target=log_status, daemon=True).start()

    app.run(host="127.0.0.1", port=5000, debug=True)
